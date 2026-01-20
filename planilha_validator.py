# validador_core.py
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
import os
import time
import os
from datetime import datetime
from openpyxl import Workbook

# Cores definidas
COR_VALIDO = PatternFill(
    start_color="00FF00", end_color="00FF00", fill_type="solid"
)  # Verde
COR_ERRO = PatternFill(
    start_color="FF0000", end_color="FF0000", fill_type="solid"
)  # Vermelho
COR_ADVERTENCIA = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)  # Amarelo
COR_DUPLICADO = PatternFill(
    start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
)  # Cinza para duplicados

# Borda para todas as células (exceto nas abas EMPRESA e RESULTADO DAS VALIDAÇÕES)
BORDA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Dicionário de Estados (para aba ESTADOS)
# Função para converter preços com pontos para formato com vírgula (SQL-friendly)
def convert_price_to_comma_format(value_str):
    """
    Converte preços com pontos para formato com vírgula (SQL-friendly)
    
    Exemplos:
    - "1.000" → ("1000,00", True) (mil reais)
    - "31.9" → ("31,90", True) (trinta e um vírgula noventa)  
    - "23,90" → ("23,90", False) (já correto)
    - "23.900000000000002" → ("23,90", True) (correção de erro de float)
    
    Retorna: (valor_convertido, foi_alterado)
    """
    original_str = str(value_str).strip()
    
    if isinstance(value_str, (int, float)):
        # Tratar erros de precisão de float primeiro
        rounded_value = round(float(value_str), 2)
        converted = f"{rounded_value:.2f}".replace(".", ",")
        return converted, (converted != original_str)
    
    value_str = original_str
    
    # Se já tem vírgula, apenas valida formato
    if "," in value_str and "." not in value_str:
        return value_str, False
    
    # Se tem ponto, precisa converter inteligentemente
    if "." in value_str:
        parts = value_str.split(".")
        
        if len(parts) == 2:
            decimal_part = parts[1]
            
            # NOVA LÓGICA: Detectar erros de precisão de float
            # Se tem muitos dígitos decimais (>4), provavelmente é erro de float
            if len(decimal_part) > 4:
                try:
                    # Tentar converter para float e arredondar para 2 casas decimais
                    float_val = float(value_str)
                    rounded_val = round(float_val, 2)
                    converted = f"{rounded_val:.2f}".replace(".", ",")
                    return converted, True
                except:
                    # Se falhar, usar lógica original
                    pass
            
            if len(decimal_part) >= 3 and len(decimal_part) <= 4:
                # Separador de milhares: "1.000" → "1000,00"
                # Só tratar como separador de milhares se for exatamente "000"
                if decimal_part == "000":
                    converted = parts[0] + decimal_part + ",00"
                    return converted, True
                else:
                    # Decimal com 3-4 dígitos, arredondar para 2
                    try:
                        float_val = float(value_str)
                        rounded_val = round(float_val, 2)
                        converted = f"{rounded_val:.2f}".replace(".", ",")
                        return converted, True
                    except:
                        converted = parts[0] + "," + decimal_part[:2]
                        return converted, True
            else:
                # Decimal mal formatado: "31.9" → "31,90"
                converted = parts[0] + "," + decimal_part.ljust(2, '0')
                return converted, True
        
        elif len(parts) == 3:
            # Formato "1.234.56" → "1234,56"
            converted = parts[0] + parts[1] + "," + parts[2]
            return converted, True
    
    # Se não tem separadores, adiciona vírgula decimal
    converted = value_str + ",00"
    return converted, (converted != original_str)


def tentar_avaliar_formula_simples(formula_str):
    """
    Tenta avaliar fórmulas muito simples como =A1*B1, =100+50, etc.
    Retorna None se não conseguir avaliar.
    """
    try:
        # Remove o sinal de igual
        expr = formula_str[1:].strip()
        
        # Só tenta avaliar se for expressão matemática simples (números e operadores)
        import re
        if re.match(r'^[\d\+\-\*\/\.\,\(\)\s]+$', expr):
            # Substitui vírgulas por pontos para cálculo
            expr = expr.replace(',', '.')
            result = eval(expr)
            return str(result)
    except:
        pass
    return None


def obter_valor_celula_com_formula(cell, nome_campo):
    """
    Tenta obter valor de célula que pode conter fórmula usando múltiplas estratégias.
    Retorna: (valor_str, mensagens_list)
    """
    if not cell or not cell.value:
        return None, []

    valor_str = str(cell.value).strip()
    mensagens = []
    
    # Se não é fórmula, retorna direto
    if not valor_str.startswith('='):
        return valor_str, mensagens
    
    # É fórmula - tentar obter valor calculado
    valor_calculado = None
    
    # Método 1: cached_value (valor já calculado pelo Excel)
    try:
        if hasattr(cell, 'cached_value') and cell.cached_value is not None:
            valor_calculado = str(cell.cached_value)
    except:
        pass
    
    # Método 2: _value (valor interno do openpyxl)
    if valor_calculado is None:
        try:
            if hasattr(cell, '_value') and cell._value is not None and not str(cell._value).startswith('='):
                valor_calculado = str(cell._value)
        except:
            pass
    
    # Método 3: tentar avaliar fórmula simples
    if valor_calculado is None:
        valor_calculado = tentar_avaliar_formula_simples(valor_str)
    
    if valor_calculado is not None:
        mensagens.append(f"Advertencia: {nome_campo} fórmula avaliada para '{valor_calculado}'")
        return valor_calculado, mensagens
    else:
        mensagens.append(f"Advertencia: {nome_campo} contém fórmula não calculada - abra o arquivo no Excel para recalcular")
        return None, mensagens


def obter_valor_celula_seguro(cell):
    """
    Função genérica para obter valor de qualquer célula, lidando com fórmulas.
    Retorna: (valor_string, sucesso_boolean)
    """
    if not cell or not cell.value:
        return "", True

    valor_str = str(cell.value).strip()
    
    # Se não é fórmula, retorna direto
    if not valor_str.startswith('='):
        return valor_str, True
    
    # É fórmula - tentar obter valor calculado
    valor_calculado = None
    
    # Método 1: cached_value (valor já calculado pelo Excel)
    try:
        if hasattr(cell, 'cached_value') and cell.cached_value is not None:
            valor_calculado = str(cell.cached_value)
    except:
        pass
    
    # Método 2: _value (valor interno do openpyxl)
    if valor_calculado is None:
        try:
            if hasattr(cell, '_value') and cell._value is not None and not str(cell._value).startswith('='):
                valor_calculado = str(cell._value)
        except:
            pass
    
    # Método 3: tentar avaliar fórmula simples (só para números)
    if valor_calculado is None:
        valor_calculado = tentar_avaliar_formula_simples(valor_str)
    
    if valor_calculado is not None:
        return valor_calculado.strip(), True
    else:
        # Retorna valor original da fórmula e marca como não calculada
        return valor_str, False


ESTADOS_BRASIL = {
    "AC": "Acre",
    "AL": "Alagoas",
    "AP": "Amapá",
    "AM": "Amazonas",
    "BA": "Bahia",
    "CE": "Ceará",
    "DF": "Distrito Federal",
    "ES": "Espírito Santo",
    "GO": "Goiás",
    "MA": "Maranhão",
    "MT": "Mato Grosso",
    "MS": "Mato Grosso do Sul",
    "MG": "Minas Gerais",
    "PA": "Pará",
    "PB": "Paraíba",
    "PR": "Paraná",
    "PE": "Pernambuco",
    "PI": "Piauí",
    "RJ": "Rio de Janeiro",
    "RN": "Rio Grande do Norte",
    "RS": "Rio Grande do Sul",
    "RO": "Rondônia",
    "RR": "Roraima",
    "SC": "Santa Catarina",
    "SP": "São Paulo",
    "SE": "Sergipe",
    "TO": "Tocantins",
}


def corrigir_campo(cell, allowed_set):
    if cell.value is None:
        return False
    valor = str(cell.value).strip() if cell.value else ""
    if valor not in allowed_set:
        cell.value = ""
        return True
    return False

def split_text(text, limit=23):
    """
    Divide um texto em duas partes, de forma que a primeira parte tenha até 'limit' caracteres,
    sem cortar palavras no meio. Se o texto for menor ou igual a 'limit', retorna (texto, "").
    Caso contrário, corta no último espaço antes de atingir o limite.
    """
    text = text.strip()
    if len(text) <= limit:
        return text, ""
    pos = text.rfind(" ", 0, limit)
    if pos == -1:
        # Se não houver espaço, corta exatamente no limite
        return text[:limit], text[limit:].strip()
    return text[:pos], text[pos:].strip()



class PlanilhaValidator:
    @staticmethod
    def get_valor_string(cell):
        """
        Converte valor de célula para string de forma otimizada.
        Evita conversões desnecessárias e já faz strip.
        """
        if cell is None or cell.value is None:
            return ""
        if isinstance(cell.value, str):
            return cell.value.strip()
        return str(cell.value).strip()

    def __init__(self, arquivo, progress_callback=None):
        """
        Args:
            arquivo: Caminho do arquivo Excel
            progress_callback: Função opcional callback(percentual, mensagem) para reportar progresso
        """
        self.arquivo = arquivo
        self.progress_callback = progress_callback

        # Carregar workbook original (preservado)
        self.wb_original = load_workbook(arquivo)
        
        # Carregar workbook apenas com valores para validação
        try:
            self.wb = load_workbook(arquivo, data_only=True)
            print("✅ Workbook carregado com data_only=True - fórmulas convertidas automaticamente")
        except Exception as e:
            print(f"⚠️ Falha ao carregar com data_only=True: {e}")
            # Fallback: usar workbook original e converter manualmente
            self.wb = load_workbook(arquivo)
            self.converter_formulas_para_valores()
        
        self.resultados_validacao = {}  # resumo por aba

        # Inicializa a variável de tempo estimado
        self.tempo_estimado_validacao = None
        
        # Cache extremamente rápido para validação de PathFotografia
        self.cache_arquivos_imagem = None
        self._carregar_cache_imagens()
        self.tempo_inicio = None

        # Dados da aba EMPRESA (para validações cruzadas)
        self.emp_nome = None
        self.emp_cod_tipo = None  # "N" ou "A"
        self.emp_cod_tamanho = None  # valor de C8
        self.emp_cod_aux = None  # "X", "N" ou "A"
        self.emp_cod_aux_tamanho = None  # valor de C11

        self.filial_cod_list = []  # do sheet FILIAL
        self.repr_cod_list = []  # Representantes
        self.pagto_cod_list = []  # CodCondPagamento (aba PAGTO)
        self.transp_dict = {}  # {CodTransportadora: Transportadora}
        self.familia_cod_list = []  # CodFamilia (aba FAMILIAS)
        self.estilo_cod_list = []  # CodEstilo (aba ESTILOS)

    def _carregar_cache_imagens(self):
        """
        Carrega cache de arquivos de imagem para validação extremamente rápida.
        Uma única leitura do diretório para milhares de verificações O(1).
        """
        import os
        caminho_imagens = r"C:\Users\Public\Documents\SRPP\Imagens_Produto"
        
        try:
            if os.path.exists(caminho_imagens):
                # Set é extremamente rápido para lookup - O(1)
                arquivos = os.listdir(caminho_imagens)
                self.cache_arquivos_imagem = set(arquivos)
                print(f"✅ Cache de imagens carregado: {len(arquivos)} arquivos")
            else:
                self.cache_arquivos_imagem = set()
                print("⚠️ Diretório de imagens não encontrado - validação PathFotografia desabilitada")
        except Exception as e:
            self.cache_arquivos_imagem = set()
            print(f"⚠️ Erro ao carregar cache de imagens: {e}")

    def _reportar_progresso(self, percentual, mensagem):
        """Reporta progresso via callback se disponível."""
        if self.progress_callback:
            self.progress_callback(percentual, mensagem)

    def limpar_espacos(self):
        """
        Remove espaços à esquerda e à direita dos dados das abas especificadas,
        exceto nos cabeçalhos e na coluna "RESULTADO".
        """
        abas_processar = ['FILIAL', 'REPR', 'PAGTO', 'PAGTOFILIAL', 'TRANSP',
                          'ESTADOS', 'CLIENTES', 'FAMILIAS', 'ESTILOS', 'PRODUTOS']
        
        for sheet_name in self.wb.sheetnames:
            if sheet_name.upper() in abas_processar:
                ws = self.wb[sheet_name]
                # Lê os cabeçalhos da primeira linha
                headers = {}
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col).value
                    if header_value:
                        headers[col] = str(header_value).strip().upper()
                # Itera sobre as linhas a partir da segunda (ignorando o cabeçalho)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if isinstance(cell.value, str):
                            header = headers.get(cell.column, "")
                            if header != "RESULTADO":
                                cell.value = cell.value.strip()   

    def tentar_recalcular_formulas(self, sheet):
        """
        Tenta recalcular fórmulas da planilha. 
        Retorna número de fórmulas não calculadas.
        """
        formulas_nao_calculadas = 0
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # célula com fórmula
                        try:
                            # Força recálculo tentando acessar o valor
                            _ = cell.value
                            if not hasattr(cell, 'cached_value') or cell.cached_value is None:
                                formulas_nao_calculadas += 1
                        except:
                            formulas_nao_calculadas += 1
        except:
            pass
        return formulas_nao_calculadas

    def converter_formulas_para_valores(self):
        """
        Converte todas as fórmulas em valores fixos nas abas especificadas.
        Preserva EMPRESA, converte todas as outras.
        """
        abas_para_converter = ["FILIAL", "REPR", "PAGTO", "PAGTOFILIAL", "TRANSP", 
                              "ESTADOS", "CLIENTES", "FAMILIAS", "ESTILOS", "PRODUTOS"]
        
        total_convertidas = 0
        total_falharam = 0
        
        for sheet_name in self.wb.sheetnames:
            if sheet_name.upper() in [aba.upper() for aba in abas_para_converter]:
                sheet = self.wb[sheet_name]
                convertidas, falharam = self._converter_formulas_aba(sheet, sheet_name)
                total_convertidas += convertidas
                total_falharam += falharam
        
        if total_convertidas > 0:
            print(f"✅ {total_convertidas} fórmulas convertidas para valores")
        if total_falharam > 0:
            print(f"⚠️ {total_falharam} fórmulas não puderam ser convertidas")

    def _converter_formulas_aba(self, sheet, sheet_name):
        """Converte fórmulas de uma aba específica para valores"""
        convertidas = 0
        falharam = 0
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # É fórmula
                    valor_original = cell.value
                    try:
                        # Método 1: usar cached_value se disponível
                        if hasattr(cell, 'cached_value') and cell.cached_value is not None:
                            # Se cached_value é float, formatar com 2 casas decimais para evitar erro de precisão
                            if isinstance(cell.cached_value, float):
                                cell.value = f"{cell.cached_value:.2f}"
                            else:
                                cell.value = cell.cached_value
                            convertidas += 1
                            continue
                        
                        # Método 2: usar _value se disponível
                        if hasattr(cell, '_value') and cell._value is not None and not str(cell._value).startswith('='):
                            cell.value = cell._value
                            convertidas += 1
                            continue
                        
                        # Método 3: tentar avaliar fórmula simples
                        if valor_original and str(valor_original).startswith('='):
                            resultado = tentar_avaliar_formula_simples(str(valor_original))
                            if resultado is not None:
                                cell.value = resultado
                                convertidas += 1
                                continue
                        
                        # Se chegou aqui, não conseguiu converter
                        falharam += 1
                        
                    except Exception as e:
                        # Em caso de erro, manter fórmula original
                        falharam += 1
        
        return convertidas, falharam

    def estimar_tempo_validacao(self):
        """
        Estima o tempo total de execução das funções de validação.
        Uma abordagem é medir individualmente algumas funções (ou usar um valor fixo se a medição não for viável).
        """
        # Exemplo: vamos simular a medição de duas funções de validação:
        tempos = {}
        
        inicio = time.time()
        # Suponha que validar_EMPRESA seja rápido
        self.validar_EMPRESA()  # Essa função não deve ter efeitos colaterais
        tempos['validar_EMPRESA'] = time.time() - inicio

        # Simule outra etapa (você pode repetir para as demais funções)
        inicio = time.time()
        self.validar_FILIAL()
        tempos['validar_FILIAL'] = time.time() - inicio

        # Some os tempos medidos – ou use um valor fixo se preferir
        total_time = sum(tempos.values())
        if total_time == 0:
            total_time = 30  # valor padrão em segundos
        self.tempo_estimado_validacao = total_time
        return total_time
    
    

    def iniciar_tempo_processamento(self):
        """
        Armazena o timestamp do início do processamento da validação.
        """
        self.tempo_inicio = time.time()
        print(
            f"Processamento iniciado em {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(self.tempo_inicio))}"
        )

        funcoes_validacao = [
            self.validar_EMPRESA,
            self.pre_validar_filial,
            self.validar_FILIAL,
            self.validar_REPR,
            self.validar_PAGTO,
            self.validar_PAGTOFILIAL,
            self.validar_TRANSP,
            self.validar_ESTADOS,
            self.validar_CLIENTES,
            self.validar_FAMILIAS,
            self.validar_ESTILOS,
            self.validar_PRODUTOS,
        ]

        tempos_execucao = []

        for func in funcoes_validacao:
            inicio = time.time()
            func()  # Executa a validação sem salvar resultados
            fim = time.time()
            tempos_execucao.append(fim - inicio)

        # Calcula tempo total estimado
        self.tempo_estimado_validacao = sum(tempos_execucao)

        print(
            f"Tempo estimado de validação: {self.tempo_estimado_validacao:.2f} segundos"
        )

    def limpar_planilha(self):
        # 1. Remove a aba "RESULTADO DAS VALIDAÇÕES", se existir.
        if "RESULTADO DAS VALIDAÇÕES" in self.wb.sheetnames:
            del self.wb["RESULTADO DAS VALIDAÇÕES"]

        # 2. Para cada aba do workbook:
        for sheet in self.wb.worksheets:
            # Obtenha o mapeamento de cabeçalhos (retorna um dicionário {nome: índice})
            header = self.get_header_map(sheet)

            # Encontra todas as colunas cujo cabeçalho seja "RESULTADO"
            indices_resultado = [
                idx for key, idx in header.items() if key.strip().upper() == "RESULTADO"
            ]
            # Exclua as colunas em ordem decrescente (para não alterar os índices das demais colunas)
            for idx in sorted(indices_resultado, reverse=True):
                sheet.delete_cols(
                    idx + 1
                )  # openpyxl trabalha com índices 1-base para exclusão

            # Se a aba for "PRODUTOS", também remova a coluna "duplicados"
            if sheet.title.upper() == "PRODUTOS":
                header = self.get_header_map(
                    sheet
                )  # Atualize o header, pois ele pode ter mudado
                indices_duplicados = [
                    idx
                    for key, idx in header.items()
                    if key.strip().lower() == "duplicados"
                ]
                for idx in sorted(indices_duplicados, reverse=True):
                    sheet.delete_cols(idx + 1)

    def converter_tudo_para_texto(self):
        for aba in self.wb.worksheets:
            for row in aba.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.value = str(cell.value)

    def gerar_status_por_aba(
        self,
        nome_aba,
        total_linhas,
        linhas_validas,
        linhas_advertencias=0,
        linhas_erros=0,
    ):
        self.resultados_validacao[nome_aba] = {
            "lidas": total_linhas,
            "validas": linhas_validas,
            "advertencias": linhas_advertencias,
            "erros": linhas_erros,
        }

    def gerar_relatorio_final(self):
        ws = self.wb.create_sheet("RESULTADO DAS VALIDAÇÕES", 0)
        headers = ["Planilha", "Mensagem"]
        ws.append(headers)

        # Estilos do cabeçalho: Fundo #00CCFF, Fonte Calibri 11 negrito
        header_fill = PatternFill(
            start_color="00CCFF", end_color="00CCFF", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        # Criar bordas para todas as células
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        abas_relatorio = [
            "EMPRESA",
            "FILIAL",
            "REPR",
            "PAGTO",
            "PAGTOFILIAL",
            "TRANSP",
            "ESTADOS",
            "CLIENTES",
            "FAMILIAS",
            "ESTILOS",
            "PRODUTOS",
        ]

        row_idx = 2
        for aba_nome in abas_relatorio:
            if aba_nome in self.resultados_validacao:
                dados = self.resultados_validacao[aba_nome]
                mensagem = (
                    f"Linhas Lidas: {dados['lidas']} | Válidas: {dados['validas']} | "
                    f"Advertências: {dados.get('advertencias', 0)} | Erros: {dados['erros']}"
                )
            else:
                mensagem = "Aba não encontrada ou não preenchida"

            # Criar link para a aba correspondente
            link_formula = f'=HYPERLINK("#\'{aba_nome}\'!A1", "{aba_nome}")'
            msg_formula = f'=HYPERLINK("#\'{aba_nome}\'!A1", "{mensagem}")'

            # Adicionar linha ao relatório
            ws.append(["", ""])  # Criar células vazias para preenchimento posterior

            # Inserir link na coluna "Planilha"
            ws.cell(row=row_idx, column=1, value=link_formula)
            ws.cell(row=row_idx, column=2, value=msg_formula)

            # Determinar a cor de fundo da linha
            if aba_nome in self.resultados_validacao:
                erros = int(self.resultados_validacao[aba_nome].get("erros", 0))
                advertencias = int(
                    self.resultados_validacao[aba_nome].get("advertencias", 0)
                )
                if erros > 0:
                    fill = COR_ERRO
                elif advertencias > 0:
                    fill = COR_ADVERTENCIA
                else:
                    fill = COR_VALIDO
            else:
                fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )

            # Aplicar estilos para cada célula na linha
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.border = thin_border  # Adiciona bordas

            # Ajustar altura da linha (exceto cabeçalho)
            ws.row_dimensions[row_idx].height = 17
            row_idx += 1

        # Ajustar largura das colunas automaticamente
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # Obtém a letra da coluna
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = (
                max_length * 0.8
            )  # Ajuste extra para espaçamento

            # Ativar filtro automático
            ws.auto_filter.ref = ws.dimensions

    def obter_result_col(self, sheet):
        for cell in sheet[1]:  # Percorre a primeira linha (cabeçalhos)
            if cell.value and str(cell.value).strip() == "RESULTADO":
                return cell.column
        return sheet.max_column + 1


    def get_header_map(self, sheet):
        header = [cell.value for cell in sheet[1]]
        return {name: idx for idx, name in enumerate(header) if name is not None}

    def get_mandatory_cell(self, row, header, field):
        idx = header.get(field)
        if idx is None:
            return None
        return row[idx]

    def escrever_resultado_linha(
        self, sheet, row_num, result_message, result_col, fill=None
    ):
        cell = sheet.cell(row=row_num, column=result_col, value=result_message)
        if fill:
            cell.fill = fill

    def determinar_fill_resultado(self, mensagens):
        # Se alguma mensagem indicar erro, retorna COR_ERRO
        if any(
            "inválido" in m.lower()
            or "duplicado" in m.lower()
            or "ausente" in m.lower()
            or "inexistente" in m.lower()
            for m in mensagens
        ):
            return COR_ERRO
        # Se não houver erro, mas houver advertência, retorna COR_ADVERTENCIA
        elif any("advertencia" in m.lower() for m in mensagens):
            return COR_ADVERTENCIA
        else:
            return COR_VALIDO

    def aplicar_borda(self, sheet):
        if sheet.title.upper() in ["EMPRESA", "RESULTADO DAS VALIDAÇÕES"]:
            return
        sheet.protection.sheet = False
        # OTIMIZAÇÃO: Aplicar borda apenas nas células com dados
        max_row = sheet.max_row
        max_col = sheet.max_column
        if max_row <= 1 or max_col <= 0:
            return
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    cell.border = BORDA
        sheet.auto_filter.ref = sheet.dimensions

    def excluir_linhas_duplicadas_produtos(self, sheet, header):
        ignore_cols = set()
        for key, idx in header.items():
            if key.lower() in ["duplicados", "resultado"]:
                ignore_cols.add(idx)

        # OTIMIZAÇÃO: Verificação rápida de linha vazia usando CodProduto
        idx_codproduto = header.get("CodProduto")

        seen = {}
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Verificação rápida de linha vazia
            if idx_codproduto is not None:
                first_val = row[idx_codproduto].value
                if first_val is None or (isinstance(first_val, str) and not first_val.strip()):
                    continue
            elif row[0].value is None:
                continue

            # Criar tupla apenas com valores relevantes (otimizado)
            row_tuple = tuple(
                str(cell.value).strip() if cell.value is not None else ""
                for idx, cell in enumerate(row)
                if idx not in ignore_cols
            )
            if row_tuple in seen:
                rows_to_delete.append(row[0].row)
            else:
                seen[row_tuple] = row[0].row

        # Deletar em ordem reversa para não afetar índices
        for r in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(r)



    @staticmethod
    def corrigir_cabecalho(sheet, expected):
        """
        Corrige os cabeçalhos da planilha comparando com a lista 'expected'.
        Se o valor na célula da primeira linha não for exatamente o esperado
        (ignorando espaços e case), ele é atualizado.
        Retorna: (mapping, corrections)
          mapping: dicionário {nome_corrigido: índice} (base 0)
          corrections: lista de strings com as correções efetuadas.
        """
        corrections = []
        new_headers = []
        for i in range(1, len(expected) + 1):
            cell = sheet.cell(row=1, column=i)
            current = cell.value if cell.value is not None else ""
            # Se o valor atual não corresponder exatamente ao esperado, corrige-o.
            if str(current).strip().lower() != expected[i - 1].lower():
                corrections.append(f"'{current}' foi alterado para '{expected[i - 1]}'")
                cell.value = expected[i - 1]
                new_headers.append(expected[i - 1])
            else:
                new_headers.append(current)
        mapping = {h: i for i, h in enumerate(new_headers)}
        return mapping, corrections

    def pre_validar_filial(self):
        filial_sheet = self.wb["FILIAL"]
        data_found = False
        for row in filial_sheet.iter_rows(min_row=2):
            if any(cell.value and str(cell.value).strip() != "" for cell in row):
                data_found = True
                break

        if not data_found:
            header = self.get_header_map(filial_sheet)
            if "CodFilial" in header and "Filial" in header:
                cod_index = header["CodFilial"]  # índice 0-based
                filial_index = header["Filial"]

                # Insere valores padrão para CodFilial e Filial
                filial_sheet.cell(row=2, column=cod_index + 1, value="1")
                filial_sheet.cell(
                    row=2,
                    column=filial_index + 1,
                    value=self.emp_nome if self.emp_nome else "",
                )

            # Nova regra: se C2 estiver vazio, copia o valor de C39 da aba EMPRESA
            # e se D2 estiver vazio, copia o valor de C40 da aba EMPRESA.
            # C2 é coluna 3 e D2 é coluna 4.
            empresa_sheet = self.wb["EMPRESA"]
            valor_C39 = empresa_sheet.cell(row=39, column=3).value  # C39
            valor_C40 = empresa_sheet.cell(row=40, column=3).value  # C40

            cell_C2 = filial_sheet.cell(row=2, column=3)
            if not (cell_C2.value and str(cell_C2.value).strip()):
                cell_C2.value = valor_C39
                cell_C2.fill = COR_ADVERTENCIA

            cell_D2 = filial_sheet.cell(row=2, column=4)
            if not (cell_D2.value and str(cell_D2.value).strip()):
                cell_D2.value = valor_C40
                cell_D2.fill = COR_ADVERTENCIA

            # Cria a coluna "RESULTADO" se ainda não existir, e insere a mensagem de correção
            result_col = filial_sheet.max_column + 1
            if not filial_sheet.cell(row=1, column=result_col).value:
                header_result = filial_sheet.cell(
                    row=1, column=result_col, value="RESULTADO"
                )
                header_result.fill = PatternFill(
                    start_color="000000", end_color="000000", fill_type="solid"
                )
                header_result.font = Font(color="FFFFFF", bold=True)

            filial_sheet.cell(
                row=2,
                column=result_col,
                value="Advertencias, CodFilial corrigido automaticamente",
            ).fill = COR_ADVERTENCIA

    # --- VALIDAÇÕES das demais abas (mantidas) ---
    def validar_EMPRESA(self):
        if "EMPRESA" not in self.wb.sheetnames:
            return "Erro: A aba EMPRESA não foi encontrada!"
        sheet = self.wb["EMPRESA"]
        total_linhas = 1
        linhas_validas = 1
        linhas_erros = 0
        msgs = []
        cell = sheet["C5"]
        if not cell.value:
            cell.fill = COR_ERRO
            msgs.append("Nome da empresa ausente")
            linhas_erros += 1
        else:
            cell.fill = COR_VALIDO
            self.emp_nome = cell.value
        cell = sheet["C7"]
        if cell.value not in ["N=Numérico", "A=Alfanumérico"]:
            cell.fill = COR_ERRO
            msgs.append("Tipo do código inválido em C7")
            linhas_erros += 1
        else:
            cell.fill = COR_VALIDO
            self.emp_cod_tipo = cell.value.split("=")[0]
        cell = sheet["C8"]
        try:
            tamanho = int(cell.value)
            if 4 <= tamanho <= 20:
                cell.fill = COR_VALIDO
                self.emp_cod_tamanho = tamanho
            else:
                cell.fill = COR_ERRO
                msgs.append("Tamanho do código principal fora do intervalo (4-20)")
                linhas_erros += 1
        except:
            cell.fill = COR_ERRO
            msgs.append("Tamanho do código principal não numérico")
            linhas_erros += 1
        cell = sheet["C10"]
        if cell.value not in ["X=Não Usado", "N=Numérico", "A=Alfanumérico"]:
            cell.fill = COR_ERRO
            msgs.append("Tipo do código auxiliar inválido em C10")
            linhas_erros += 1
        else:
            cell.fill = COR_VALIDO
            self.emp_cod_aux = cell.value.split("=")[0]
        
        # Validar tamanho do código auxiliar (C11) - só se não for "X=Não Usado"
        if self.emp_cod_aux != "X":
            cell = sheet["C11"]
            try:
                tamanho_aux = int(cell.value)
                if 4 <= tamanho_aux <= 20:
                    cell.fill = COR_VALIDO
                    self.emp_cod_aux_tamanho = tamanho_aux
                else:
                    cell.fill = COR_ERRO
                    msgs.append("Tamanho do código auxiliar fora do intervalo (4-20)")
                    linhas_erros += 1
            except:
                cell.fill = COR_ERRO
                msgs.append("Tamanho do código auxiliar não numérico")
                linhas_erros += 1
        
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
        self.escrever_resultado_linha(
            sheet, 5, "; ".join(msgs), result_col, self.determinar_fill_resultado(msgs)
        )
        self.gerar_status_por_aba("EMPRESA", total_linhas, linhas_validas, 0, linhas_erros)
        return None

    def validar_FILIAL(self):
        if "FILIAL" not in self.wb.sheetnames:
            return "Erro: A aba FILIAL não foi encontrada!"

        sheet = self.wb["FILIAL"]


        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodFilial", "Filial", "TituloAdicional1", "TituloAdicional2", "Logotipo"]
        
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)


        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            # Garanta que a palavra "advertencia" esteja em minúsculas para que a função determinar_fill_resultado a reconheça.
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                            " estavam com nome errado, o correto é " +
                            ", ".join(correctos))



        total_linhas = 0
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        seen_codfilial = set()
        seen_filial = set()

        # Obtém a coluna "RESULTADO" (sem duplicar)
        result_col = self.obter_result_col(sheet)

        # Cria o cabeçalho "RESULTADO" se não existir
        if not sheet.cell(row=1, column=result_col).value:
            header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
            header_result.fill = PatternFill(
                start_color="000000", end_color="000000", fill_type="solid"
            )
            header_result.font = Font(color="FFFFFF", bold=True)

        

        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue

            total_linhas += 1
            mensagens = []
            advertencia_presente = False  # Flag para saber se há advertências
            
            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)


            # Nova regra: para a linha 2, verificar C2 e D2
            if row[0].row == 2:
                cell_C2 = sheet.cell(row=2, column=3)  # C2
                cell_D2 = sheet.cell(row=2, column=4)  # D2
                empresa_sheet = self.wb["EMPRESA"]
                if not (cell_C2.value and str(cell_C2.value).strip()):
                    valor_C39 = empresa_sheet.cell(row=39, column=3).value
                    cell_C2.value = valor_C39
                    cell_C2.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, C2 corrigido automaticamente")
                    advertencia_presente = True
                if not (cell_D2.value and str(cell_D2.value).strip()):
                    valor_C40 = empresa_sheet.cell(row=40, column=3).value
                    cell_D2.value = valor_C40
                    cell_D2.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, D2 corrigido automaticamente")
                    advertencia_presente = True

            # Validação da coluna CodFilial
            cod_cell = self.get_mandatory_cell(row, header, "CodFilial")
            filial_cell = self.get_mandatory_cell(row, header, "Filial")

            if cod_cell is None:
                mensagens.append("CodFilial ausente")
            else:
                valor = self.get_valor_string(cod_cell)
                if valor and len(valor) > 40:
                    cod_cell.fill = COR_ADVERTENCIA
                    mensagens.append(
                        "Advertencia, 'CodFilial' excedeu o limite de caracteres"
                    )
                    advertencia_presente = True
                if not valor:
                    cod_cell.value = "1"
                    valor = "1"  # Atualiza a variável com o novo valor
                    cod_cell.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, CodFilial corrigido automaticamente")
                    advertencia_presente = True
                elif not valor.isdigit() or not (1 <= int(valor) <= 999999):
                    cod_cell.fill = COR_ERRO
                    mensagens.append("CodFilial inválido")
                else:
                    cod_cell.fill = COR_VALIDO
                if valor in seen_codfilial:
                    cod_cell.fill = COR_ERRO
                    mensagens.append("CodFilial duplicado")
                else:
                    seen_codfilial.add(valor)
                    self.filial_cod_list.append(valor)

            # Validação da coluna Filial
            if filial_cell is None:
                mensagens.append("Filial ausente")
            else:
                filial_val = self.get_valor_string(filial_cell)
                if filial_val and len(filial_val) > 40:
                    filial_cell.fill = COR_ADVERTENCIA
                    mensagens.append(
                        "Advertencia, 'Filial' excedeu o limite de caracteres"
                    )
                    advertencia_presente = True
                if not filial_val:
                    if self.emp_nome:
                        filial_cell.value = self.emp_nome
                        filial_cell.fill = COR_ADVERTENCIA
                        mensagens.append(
                            "Advertencia, Filial corrigido automaticamente"
                        )
                        advertencia_presente = True
                    else:
                        filial_cell.fill = COR_ERRO
                        mensagens.append("Filial ausente e sem nome da empresa")
                else:
                    filial_cell.fill = COR_VALIDO
                if filial_val in seen_filial:
                    filial_cell.fill = COR_ERRO
                    mensagens.append("Filial duplicada")
                else:
                    seen_filial.add(filial_val)

            # Define a cor de preenchimento para a linha com base nas advertências
            if advertencia_presente:
                resultado_fill = COR_ADVERTENCIA
            else:
                resultado_fill = self.determinar_fill_resultado(mensagens)

            # Aplica a cor em todas as células da linha, exceto a coluna "RESULTADO"
            for cell in row:
                if cell.column != result_col:
                    cell.fill = resultado_fill

            # Atualiza os contadores
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)
                
        # Gera o resumo da validação para a aba FILIAL
        self.gerar_status_por_aba(
            "FILIAL", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )

        # Reaplica a cor preta no cabeçalho da coluna "RESULTADO" para evitar perda da formatação
        header_result = sheet.cell(row=1, column=result_col)
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)

        # Aplica bordas finais na aba
        self.aplicar_borda(sheet)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2
        return None

    def validar_REPR(self):
        if "REPR" not in self.wb.sheetnames:
            return "Erro", "A aba REPR não foi encontrada!"

        sheet = self.wb["REPR"]

        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodRepresentante", "Representante"]
        
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            # Garanta que a palavra "advertencia" esteja em minúsculas para que a função determinar_fill_resultado a reconheça.
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                            " estavam com nome errado, o correto é " +
                            ", ".join(correctos))


        total_linhas = 0
        linhas_validas = 0
        linhas_advertencias = 0
        linhas_erros = 0
        seen_cod = set()
        seen_repr = set()

        # Cria a coluna RESULTADO
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)


        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []
        
            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)


            # Validação do CodRepresentante
            cell = self.get_mandatory_cell(row, header, "CodRepresentante")
            if cell is None:
                mensagens.append("CodRepresentante ausente")
            else:
                valor = self.get_valor_string(cell)
                if not valor or not valor.isdigit() or not (1 <= int(valor) <= 32767):
                    cell.fill = COR_ERRO
                    mensagens.append("CodRepresentante inválido")
                else:
                    cell.fill = COR_VALIDO
                    if valor in seen_cod:
                        cell.fill = COR_ERRO
                        mensagens.append("CodRepresentante duplicado")
                    else:
                        seen_cod.add(valor)
                        self.repr_cod_list.append(valor)

            # Validação do Representante
            cell_repr = self.get_mandatory_cell(row, header, "Representante")
            if cell_repr is None:
                mensagens.append("Representante ausente")
            else:
                repr_val = self.get_valor_string(cell_repr)
                if not repr_val or len(repr_val) > 20:
                    cell_repr.fill = COR_ERRO
                    mensagens.append(
                        "Advertencia, 'Representante' excede 20 caracteres"
                    )
                else:
                    cell_repr.fill = COR_VALIDO
            if repr_val in seen_repr:
                cell_repr.fill = COR_ADVERTENCIA  
                mensagens.append("Advertencia, Representante repetido")
            else:
                seen_repr.add(repr_val)

            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)

        # Se nenhuma linha foi processada, insere uma mensagem de erro na primeira linha de dados
        if total_linhas == 0:
            cell_result = sheet.cell(
                row=2,
                column=result_col,
                value="Inválido, ao menos um representante deve ser cadastrado",
            )
            cell_result.fill = COR_ERRO
            linhas_erros += 1

        self.gerar_status_por_aba(
            "REPR", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )
        self.aplicar_borda(sheet)

        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    ax_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2

        # Ajuste final da largura da coluna RESULTADO (duplicado no código original; pode ser removido se desnecessário)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2

        return None

    def validar_PAGTO(self):
        if "PAGTO" not in self.wb.sheetnames:
            return "Erro", "A aba PAGTO não foi encontrada!"

        sheet = self.wb["PAGTO"]
        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodCondPagamento",	"CondPagamento",	"TipoCondPagamento",	"CondPagamentoPadrao",	"VlrMinimoPedido",	"VlrMinimoComEstAtual",	"VlrMinimoComEstFuturo",	"VlrMinimoComEstEsgotado",	"Desconto1",	"Desconto2",	"Desconto3"]

        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            # Garanta que a palavra "advertencia" esteja em minúsculas para que a função determinar_fill_resultado a reconheça.
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                          " estavam com nome errado, o correto é " +
                          ", ".join(correctos))


        total_linhas = 0
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        seen_cod = set()
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
       
       
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []
            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)



            cell = self.get_mandatory_cell(row, header, "CodCondPagamento")
            if cell is None:
                mensagens.append("CodCondPagamento ausente")
            else:
                valor = self.get_valor_string(cell)
                if not valor or not valor.isdigit() or not (1 <= int(valor) <= 32767):
                    cell.fill = COR_ERRO
                    mensagens.append("CodCondPagamento inválido")
                else:
                    cell.fill = COR_VALIDO
                if valor in seen_cod:
                    cell.fill = COR_ERRO
                    mensagens.append("CodCondPagamento duplicado")
                else:
                    seen_cod.add(valor)
                    self.pagto_cod_list.append(valor)
            cell_cond = self.get_mandatory_cell(row, header, "CondPagamento")
            if cell_cond is None:
                mensagens.append("CondPagamento ausente")
            else:
                cond_val = self.get_valor_string(cell_cond)
                if not cond_val or len(cond_val) > 20:
                    cell_cond.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, 'CondPagamento' excede 20 caracteres")
                else:
                    cell_cond.fill = COR_VALIDO
            cell_tipo = self.get_mandatory_cell(row, header, "TipoCondPagamento")
            if cell_tipo is None:
                mensagens.append("TipoCondPagamento ausente")
            else:
                if corrigir_campo(cell_tipo, {"N", "E", "n", "e", ""}):
                    cell_tipo.fill = COR_ERRO
                    mensagens.append("TipoCondPagamento inválido")
                else:
                    cell_tipo.fill = COR_VALIDO
            cell_padrao = self.get_mandatory_cell(row, header, "CondPagamentoPadrao")
            if cell_padrao is None:
                mensagens.append("CondPagamentoPadrao ausente")
            else:
                if corrigir_campo(cell_padrao, {"S", "s", "N", "n", ""}):
                    cell_padrao.fill = COR_ERRO
                    mensagens.append("CondPagamentoPadrao inválido")
                else:
                    cell_padrao.fill = COR_VALIDO
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)

        self.gerar_status_por_aba("PAGTO", total_linhas, linhas_validas, linhas_advertencias, linhas_erros)
        self.aplicar_borda(sheet)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2
        return None

    def validar_PAGTOFILIAL(self):
        if "PAGTOFILIAL" not in self.wb.sheetnames:
            return
        
        sheet = self.wb["PAGTOFILIAL"]

        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodCondPagamento", "CodFilial", "VlrMinimoPedido"]
        
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            # Garanta que a palavra "advertencia" esteja em minúsculas para que a função determinar_fill_resultado a reconheça.
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                              " estavam com nome errado, o correto é " +
                              ", ".join(correctos))
        total_linhas = 0
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []

            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)

            cell = self.get_mandatory_cell(row, header, "CodCondPagamento")
            if cell is None:
                mensagens.append("CodCondPagamento ausente")
            else:
                valor = self.get_valor_string(cell)
                if valor not in self.pagto_cod_list:
                    cell.fill = COR_ERRO
                    mensagens.append("CodCondPagamento inexistente na aba PAGTO")
                else:
                    cell.fill = COR_VALIDO
            cell_fil = self.get_mandatory_cell(row, header, "CodFilial")
            if cell_fil is None:
                mensagens.append("CodFilial ausente")
            else:
                valor_fil = self.get_valor_string(cell_fil)
                if valor_fil not in self.filial_cod_list:
                    cell_fil.fill = COR_ERRO
                    mensagens.append("CodFilial inexistente na aba FILIAL")
                else:
                    cell_fil.fill = COR_VALIDO
            idx = header.get("VlrMinimoPedido")
            if idx is not None:
                cell_valor = row[idx]
                if cell_valor.value:
                    try:
                        valor_str = self.get_valor_string(cell_valor)
                        valor_float = float(valor_str.replace(",", "."))
                        if not (0.00 <= valor_float <= 9999999999.99):
                            cell_valor.fill = COR_ERRO
                            mensagens.append("VlrMinimoPedido fora do intervalo")
                        else:
                            cell_valor.fill = COR_VALIDO
                            cell_valor.value = f"{valor_float:.2f}".replace(".", ",")
                    except:
                        cell_valor.fill = COR_ERRO
                        mensagens.append("VlrMinimoPedido inválido")
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)

        self.gerar_status_por_aba(
            "PAGTOFILIAL",
            total_linhas,
            linhas_validas,
            linhas_advertencias,
            linhas_erros,
        )
        self.aplicar_borda(sheet)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2

    def validar_TRANSP(self):
        if "TRANSP" not in self.wb.sheetnames:
            return
        sheet = self.wb["TRANSP"]


        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodTransportadora", "Transportadora", "TransportadoraPadrao"]
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            # Garanta que a palavra "advertencia" esteja em minúsculas para que a função determinar_fill_resultado a reconheça.
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                            " estavam com nome errado, o correto é " +
                            ", ".join(correctos))

        total_linhas = 0
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        seen_cod = set()
        seen_nome = set()
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []

            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)



            cell = self.get_mandatory_cell(row, header, "CodTransportadora")
            if cell is None:
                mensagens.append("CodTransportadora ausente")
            else:
                valor = self.get_valor_string(cell)
                if not valor or not valor.isdigit() or not (1 <= int(valor) <= 32767):
                    cell.fill = COR_ERRO
                    mensagens.append("CodTransportadora inválido")
                else:
                    cell.fill = COR_VALIDO
                if valor in seen_cod:
                    cell.fill = COR_ERRO
                    mensagens.append("CodTransportadora duplicado")
                else:
                    seen_cod.add(valor)
                    self.transp_dict[valor] = None
            cell_transp = self.get_mandatory_cell(row, header, "Transportadora")
            if cell_transp is None:
                mensagens.append("Transportadora ausente")
            else:
                nome_transp = self.get_valor_string(cell_transp)
                if not nome_transp or len(nome_transp) > 20:
                    cell_transp.fill = COR_ERRO
                    mensagens.append("Transportadora inválida ou excede 20 caracteres")
                else:
                    cell_transp.fill = COR_VALIDO
                if nome_transp in seen_nome:
                    cell_transp.fill = COR_ERRO
                    mensagens.append("Transportadora duplicada")
                else:
                    seen_nome.add(nome_transp)
                    self.transp_dict[valor] = nome_transp
            cell_pad = self.get_mandatory_cell(row, header, "TransportadoraPadrao")
            if cell_pad is None:
                mensagens.append("TransportadoraPadrao ausente")
            else:
                if corrigir_campo(cell_pad, {"S", "s", "N", "n", ""}):
                    cell_pad.fill = COR_ERRO
                    mensagens.append("TransportadoraPadrao inválido")
                else:
                    cell_pad.fill = COR_VALIDO
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)

        self.gerar_status_por_aba(
            "TRANSP", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )
        self.aplicar_borda(sheet)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2

    def validar_ESTADOS(self):
        if "ESTADOS" not in self.wb.sheetnames:
            return
        sheet = self.wb["ESTADOS"]


        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["SiglaEstado", "NomeEstado", "Padrao", "ClienteNovoTabPreco"]
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            # Garanta que a palavra "advertencia" esteja em minúsculas para que a função determinar_fill_resultado a reconheça.
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                            " estavam com nome errado, o correto é " +
                            ", ".join(correctos))
        total_linhas = 0
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []
            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)
            cell_sigla = self.get_mandatory_cell(row, header, "SiglaEstado")
            sigla = ""
            if cell_sigla is None:
                mensagens.append("SiglaEstado ausente")
            else:
                sigla = self.get_valor_string(cell_sigla)
                if sigla not in ESTADOS_BRASIL:
                    cell_sigla.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, 'SiglaEstado' inválida")
                else:
                    cell_sigla.fill = COR_VALIDO
            cell_nome = self.get_mandatory_cell(row, header, "NomeEstado")
            if cell_nome is None:
                mensagens.append("NomeEstado ausente")
            else:
                nome_estado = self.get_valor_string(cell_nome)
                esperado = ESTADOS_BRASIL.get(sigla, "")
                if sigla in ESTADOS_BRASIL and nome_estado.lower() != esperado.lower():
                    cell_nome.value = esperado
                    cell_nome.fill = COR_VALIDO
                else:
                    if len(nome_estado) > 20:
                        cell_nome.fill = COR_ADVERTENCIA
                        mensagens.append(
                            "Advertencia, 'NomeEstado' excedeu o limite de caracteres"
                        )
                    else:
                        cell_nome.fill = COR_VALIDO
            cell_pad = self.get_mandatory_cell(row, header, "Padrao")
            if cell_pad is None:
                mensagens.append("Padrao ausente")
            else:
                if corrigir_campo(cell_pad, {"1", "2"}):
                    cell_pad.fill = COR_ERRO
                    mensagens.append("Padrao inválido")
                else:
                    cell_pad.fill = COR_VALIDO
            idx = header.get("ClienteNovoTabPreco")
            if idx is not None:
                cell_cliente = row[idx]
                if cell_cliente.value:
                    try:
                        val = int(cell_cliente.value)
                        if val not in [0, 1, 2, 3]:
                            cell_cliente.fill = COR_ERRO
                            mensagens.append("ClienteNovoTabPreco deve ser 0,1,2 ou 3")
                        else:
                            cell_cliente.fill = COR_VALIDO
                    except:
                        cell_cliente.fill = COR_ERRO
                        mensagens.append("ClienteNovoTabPreco inválido")
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)

        self.gerar_status_por_aba(
            "ESTADOS", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )
        self.aplicar_borda(sheet)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2

        # Ao final do método validar_XXX, depois de aplicar bordas e ajustar colunas:

    def validar_CLIENTES(self):
        if "CLIENTES" not in self.wb.sheetnames:
            return
        sheet = self.wb["CLIENTES"]

        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodCliente", "NomeFantasia", "CodRepresentante", "RazaoSocial", "Logradouro", "Bairro", "Cidade",	"UF",	"CEP",	"CNPJCPF",	"IERG",	"Observacao",	"CodTransportadora",	"NomeTransportadora",	"PrecoTabela",	"NomeContato", "EMail", "DDD", "Telefone1",	"Telefone2"]
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                      " estavam com nome errado, o correto é " +
                      ", ".join(correctos))

        total_linhas = 0
        linhas_validas = 0
        linhas_advertencias = 0
        linhas_erros = 0
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
        
        # Rastrear CodCliente normalizados para detectar duplicatas considerando zeros à esquerda
        # Formato: {valor_normalizado: (valor_original_string, numero_linha)}
        seen_codcliente_normalized = {}
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []
            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)
            cell = self.get_mandatory_cell(row, header, "CodCliente")
            if cell is None:
                mensagens.append("CodCliente ausente")
            else:
                valor = self.get_valor_string(cell)
                if not valor or not valor.isdigit() or not (1 <= int(valor) <= 9999999):
                    cell.fill = COR_ERRO
                    mensagens.append("CodCliente inválido")
                else:
                    # Verificar duplicatas considerando zeros à esquerda
                    normalized_cod = int(valor)  # Remove zeros à esquerda
                    if normalized_cod in seen_codcliente_normalized:
                        original_cod, original_linha = seen_codcliente_normalized[normalized_cod]
                        linha_atual = row[0].row
                        
                        # Se são strings diferentes (zeros à esquerda), mostrar exemplo com linhas
                        if original_cod != valor:
                            mensagem = f"CodCliente duplicado: {valor} na linha {linha_atual} já existe como {original_cod} na linha {original_linha}"
                        else:
                            mensagem = f"CodCliente duplicado: {valor} na linha {linha_atual} já existe na linha {original_linha}"
                            
                        cell.fill = COR_ERRO
                        mensagens.append(mensagem)
                    else:
                        linha_atual = row[0].row
                        seen_codcliente_normalized[normalized_cod] = (valor, linha_atual)
                        cell.fill = COR_VALIDO
            cell_nf = self.get_mandatory_cell(row, header, "NomeFantasia")
            if cell_nf is None:
                mensagens.append("NomeFantasia ausente")
            else:
                nf = self.get_valor_string(cell_nf)
                if not nf or len(nf) > 20:
                    cell_nf.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, NomeFantasia excede 20 carcteres")
                else:
                    cell_nf.fill = COR_VALIDO
            idx = header.get("CodRepresentante")
            if idx is not None:
                cell_cr = row[idx]
                cr_val = self.get_valor_string(cell_cr)
                if cr_val == "0":
                    cell_cr.value = ""
                    cell_cr.fill = COR_VALIDO
                elif cr_val:
                    if not cr_val.isdigit() or cr_val not in self.repr_cod_list:
                        cell_cr.fill = COR_ERRO
                        mensagens.append("CodRepresentante inexistente")
                    else:
                        cell_cr.fill = COR_VALIDO
            idx = header.get("RazaoSocial")
            if idx is not None:
                cell_rs = row[idx]
                rs_val = self.get_valor_string(cell_rs)
                if rs_val and len(rs_val) > 40:
                    cell_rs.fill = COR_ERRO
                    mensagens.append("RazaoSocial excede 40 caracteres")
                else:
                    cell_rs.fill = COR_VALIDO
            idx = header.get("PrecoTabela")
            if idx is not None:
                cell_pt = row[idx]
                if cell_pt.value:
                    try:
                        val = int(cell_pt.value)
                        if val not in [0, 1, 2, 3]:
                            cell_pt.fill = COR_ERRO
                            mensagens.append("PrecoTabela deve ser 0,1,2 ou 3")
                        else:
                            cell_pt.fill = COR_VALIDO
                    except:
                        cell_pt.fill = COR_ERRO
                        mensagens.append("PrecoTabela inválido")
                else:
                    # Se estiver em branco, é válido
                    cell_pt.fill = COR_VALIDO
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)

        self.gerar_status_por_aba(
            "CLIENTES", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )
        self.aplicar_borda(sheet)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2

    def validar_FAMILIAS(self):
        if "FAMILIAS" not in self.wb.sheetnames:
            return
        sheet = self.wb["FAMILIAS"]


        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodFamilia", "Familia", "MultiploFamilia", "MinimoFamilia", "DescontoFamilia"]
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                      " estavam com nome errado, o correto é " +
                      ", ".join(correctos))
        total_linhas = 0    
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []
            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)

                
            cell = self.get_mandatory_cell(row, header, "CodFamilia")
            if cell is None:
                mensagens.append("CodFamilia ausente")
            else:
                valor = self.get_valor_string(cell)
                if not valor or not valor.isdigit() or not (1 <= int(valor) <= 999999):
                    cell.fill = COR_ERRO
                    mensagens.append("CodFamilia inválido")
                else:
                    cell.fill = COR_VALIDO
                    if valor not in self.familia_cod_list:
                        self.familia_cod_list.append(valor)
            cell_fam = self.get_mandatory_cell(row, header, "Familia")
            if cell_fam is None:
                mensagens.append("Familia ausente")
            else:
                fam_val = self.get_valor_string(cell_fam)
                if not fam_val or len(fam_val) > 45:
                    cell_fam.fill = COR_ERRO
                    mensagens.append("Familia excede 45 caracteres")
                else:
                    cell_fam.fill = COR_VALIDO
            for campo in ["MultiploFamilia", "MinimoFamilia"]:
                if campo in header:
                    idx = header.get(campo)
                    cell_val = row[idx]
                    if cell_val.value:
                        try:
                            num = int(cell_val.value)
                            if not (1 <= num <= 999999):
                                cell_val.fill = COR_ERRO
                                mensagens.append(f"{campo} fora do intervalo")
                            else:
                                cell_val.fill = COR_VALIDO
                        except:
                            cell_val.fill = COR_ERRO
                            mensagens.append(f"{campo} inválido")
            idx = header.get("DescontoFamilia")
            if idx is not None:
                cell_desc = row[idx]
                if cell_desc.value:
                    try:
                        desc_str = self.get_valor_string(cell_desc)
                        desc_val = float(desc_str.replace(",", "."))
                        if not (0.00 <= desc_val <= 99.99):
                            cell_desc.fill = COR_ERRO
                            mensagens.append("DescontoFamilia fora do intervalo")
                        else:
                            cell_desc.fill = COR_VALIDO
                            cell_desc.value = f"{desc_val:.2f}".replace(".", ",")
                    except:
                        cell_desc.fill = COR_ERRO
                        mensagens.append("DescontoFamilia inválido")
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0
            

            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)


        self.gerar_status_por_aba(
            "FAMILIAS", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )
        self.aplicar_borda(sheet)

    def validar_ESTILOS(self):
        if "ESTILOS" not in self.wb.sheetnames:
            return
        sheet = self.wb["ESTILOS"]


        # Define a lista dos cabeçalhos esperados para essa aba
        expected = ["CodEstilo", "Estilo"]
        # Chama o método para normalizar (corrigir) os cabeçalhos
        header, corrections = self.corrigir_cabecalho(sheet, expected)

        # Se houver correções, monta a mensagem informando apenas os cabeçalhos que foram alterados
        header_warning = ""
        if corrections:
            correctos = []
            for corr in corrections:
                parts = corr.split(" para ")
                if len(parts) == 2:
                    correct_header = parts[1].strip().strip("'")
                    correctos.append(correct_header)
                else:
                    correctos.append(corr)
            header_warning = ("Advertencia: " + ", ".join(corrections) +
                      " estavam com nome errado, o correto é " +
                      ", ".join(correctos))
        total_linhas = 0
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        result_col = sheet.max_column + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        header_result.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            total_linhas += 1
            mensagens = []
            # Se houver correções no cabeçalho, adiciona essa advertência para cada linha
            if header_warning:
                mensagens.append(header_warning)
            cell = self.get_mandatory_cell(row, header, "CodEstilo")
            if cell is None:
                mensagens.append("CodEstilo ausente")
            else:
                valor = self.get_valor_string(cell)
                if not valor or not valor.isdigit() or not (1 <= int(valor) <= 999999):
                    cell.fill = COR_ERRO
                    mensagens.append("CodEstilo inválido")
                else:
                    cell.fill = COR_VALIDO
                    if valor not in self.estilo_cod_list:
                        self.estilo_cod_list.append(valor)
            cell_est = self.get_mandatory_cell(row, header, "Estilo")
            if cell_est is None:
                mensagens.append("Estilo ausente")
            else:
                est_val = self.get_valor_string(cell_est)
                if not est_val or len(est_val) > 45:
                    cell_est.fill = COR_ERRO
                    mensagens.append("Estilo excede 45 caracteres")
                else:
                    cell_est.fill = COR_VALIDO
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0

            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)
        self.gerar_status_por_aba(
            "ESTILOS", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )
        self.aplicar_borda(sheet)

    def validar_PRODUTOS(self, progress_base=None, progress_range=None):
        """
        Valida a aba PRODUTOS.

        Args:
            progress_base: Percentual base para reportar progresso (ex: 77)
            progress_range: Range de percentual disponível para esta validação (ex: 15 significa 77-92%)
        """
        if "PRODUTOS" not in self.wb.sheetnames:
            return "Erro", "A aba PRODUTOS não foi encontrada!"

        sheet = self.wb["PRODUTOS"]
        # Obtém o header atual da planilha, sem forçar uma ordem específica
        header = self.get_header_map(sheet)
        header_warning = ""

        # Excluir linhas duplicadas (linhas idênticas)
        self.excluir_linhas_duplicadas_produtos(sheet, header)

        # PASSADA 1: Contagem de duplicados + limpeza de zeros (OTIMIZADO)
        seen_codproduto = {}
        seen_codaux = {}
        idx_codproduto = header.get("CodProduto")
        idx_codaux = header.get("CodAuxiliarProduto")
        total_linhas_planilha = 0  # Contar linhas válidas reais

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Verificação rápida de linha vazia
            if idx_codproduto is not None:
                first_val = row[idx_codproduto].value
                if first_val is None or (isinstance(first_val, str) and first_val.strip() == ""):
                    continue

            total_linhas_planilha += 1  # Contar apenas linhas com dados

            # Limpar zeros em todas as células da linha
            for cell in row:
                if cell.value == 0 or (cell.value is not None and str(cell.value).strip() == "0"):
                    cell.value = ""

            # Contar duplicados
            if idx_codproduto is not None and row[idx_codproduto].value:
                cp_val = str(row[idx_codproduto].value).strip()
                if cp_val:
                    seen_codproduto[cp_val] = seen_codproduto.get(cp_val, 0) + 1

            if idx_codaux is not None and row[idx_codaux].value:
                aux_val = str(row[idx_codaux].value).strip()
                if aux_val:
                    seen_codaux[aux_val] = seen_codaux.get(aux_val, 0) + 1

        # Verifica se há duplicatas
        any_duplicates = any(v > 1 for v in seen_codproduto.values()) or any(v > 1 for v in seen_codaux.values())

        # Adiciona coluna "Duplicados" se necessário
        if any_duplicates and "Duplicados" not in header:
            if "CodProduto" in header:
                codproduto_col = header["CodProduto"] + 1
                sheet.insert_cols(codproduto_col)
                dup_header = sheet.cell(row=1, column=codproduto_col, value="Duplicados")
                dup_header.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                dup_header.font = Font(color="FFFFFF", bold=True)
                header = self.get_header_map(sheet)

        # Configurar coluna RESULTADO
        result_col = len(header) + 1
        header_result = sheet.cell(row=1, column=result_col, value="RESULTADO")
        header_result.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_result.font = Font(color="FFFFFF", bold=True)

        # Ocultar colunas extras
        for col in range(result_col + 1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].hidden = True

        # Configuração de progresso
        intervalo_progresso = max(100, total_linhas_planilha // 100) if total_linhas_planilha > 0 else 100
        linha_atual = 0
        ultimo_progresso_reportado = -1

        # PASSADA 2: Validação completa
        total_linhas = 0
        linhas_validas = 0
        linhas_erros = 0
        linhas_advertencias = 0
        seen_produto = {}

        # OTIMIZAÇÃO: Pré-calcular índices fora do loop (evita 79k lookups)
        idx_codproduto = header.get("CodProduto")
        idx_codaux = header.get("CodAuxiliarProduto")
        idx_produto = header.get("Produto")
        idx_codfilial = header.get("CodFilial")
        idx_codfamilia = header.get("CodFamilia")
        idx_codestilo = header.get("CodEstilo")
        idx_duplicados = header.get("Duplicados")

        # OTIMIZAÇÃO: Pré-calcular valores da empresa
        emp_cod_tipo = self.emp_cod_tipo
        emp_cod_tamanho = self.emp_cod_tamanho
        emp_cod_aux = self.emp_cod_aux
        emp_cod_aux_tamanho = self.emp_cod_aux_tamanho
        filial_cod_list = self.filial_cod_list
        filial_unica = len(filial_cod_list) == 1
        filial_valor_unico = filial_cod_list[0] if filial_unica else None
        familia_cod_set = set(self.familia_cod_list)  # Set para lookup O(1)
        estilo_cod_set = set(self.estilo_cod_list)    # Set para lookup O(1)

        # OTIMIZAÇÃO: Função inline para obter valor string
        def get_val(cell):
            if cell is None or cell.value is None:
                return ""
            v = cell.value
            return v.strip() if isinstance(v, str) else str(v).strip()

        for row in sheet.iter_rows(min_row=2):
            # Verificação rápida de linha vazia
            if idx_codproduto is not None:
                first_val = row[idx_codproduto].value
                if first_val is None or (isinstance(first_val, str) and not first_val.strip()):
                    continue
            elif all(c.value is None for c in row[:3]):
                continue

            # Reportar progresso granular
            linha_atual += 1
            if progress_base is not None and progress_range is not None and total_linhas_planilha > 0:
                if linha_atual % intervalo_progresso == 0:
                    progresso_linha = int((linha_atual / total_linhas_planilha) * 100)
                    if progresso_linha != ultimo_progresso_reportado:
                        ultimo_progresso_reportado = progresso_linha
                        percentual_atual = progress_base + int((linha_atual / total_linhas_planilha) * progress_range)
                        self._reportar_progresso(percentual_atual, f"Validando PRODUTOS... {progresso_linha}% ({linha_atual}/{total_linhas_planilha})")

            total_linhas += 1
            mensagens = []
            dup_valores = []

            if header_warning:
                mensagens.append(header_warning)

            # Validar CodProduto (usando índice direto)
            if idx_codproduto is None:
                mensagens.append("CodProduto ausente")
            else:
                cell_cp = row[idx_codproduto]
                cp_val = get_val(cell_cp)
                if emp_cod_tipo == "N":
                    if not cp_val.isdigit():
                        cell_cp.fill = COR_ERRO
                        mensagens.append("CodProduto inválido (deve ser numérico)")
                    elif len(cp_val) > emp_cod_tamanho:
                        cell_cp.fill = COR_ERRO
                        mensagens.append("CodProduto inválido (excede tamanho permitido)")
                    else:
                        cell_cp.fill = COR_VALIDO
                elif emp_cod_tipo == "A":
                    if len(cp_val) > emp_cod_tamanho:
                        cell_cp.fill = COR_ERRO
                        mensagens.append("CodProduto excede tamanho permitido")
                    else:
                        cell_cp.fill = COR_VALIDO
                if cp_val and seen_codproduto.get(cp_val, 0) > 1:
                    mensagens.append("CodProduto duplicado")
                    dup_valores.append(cp_val)

            # Validar CodAuxiliarProduto (usando índice direto)
            if idx_codaux is not None:
                aux_cell = row[idx_codaux]
                aux_val = get_val(aux_cell)
                if aux_val:
                    if emp_cod_aux == "X":
                        aux_cell.fill = COR_ERRO
                        mensagens.append("CodAuxiliarProduto não permitido (configurado como não usado)")
                    elif emp_cod_aux == "N":
                        if not aux_val.isdigit():
                            aux_cell.fill = COR_ERRO
                            mensagens.append("CodAuxiliarProduto inválido (deve ser numérico)")
                        elif emp_cod_aux_tamanho and len(aux_val) > emp_cod_aux_tamanho:
                            aux_cell.fill = COR_ERRO
                            mensagens.append("CodAuxiliarProduto inválido (excede tamanho permitido)")
                        else:
                            aux_cell.fill = COR_VALIDO
                    elif emp_cod_aux == "A":
                        if emp_cod_aux_tamanho and len(aux_val) > emp_cod_aux_tamanho:
                            aux_cell.fill = COR_ERRO
                            mensagens.append("CodAuxiliarProduto inválido (excede tamanho permitido)")
                        else:
                            aux_cell.fill = COR_VALIDO
                    if seen_codaux.get(aux_val, 0) > 1:
                        mensagens.append("CodAuxiliarProduto duplicado")
                        dup_valores.append(aux_val)
                else:
                    aux_cell.fill = COR_VALIDO

            # Validar Produto (usando índice direto)
            if idx_produto is None:
                mensagens.append("Produto ausente")
            else:
                cell_prod = row[idx_produto]
                prod_val = get_val(cell_prod)
                if not prod_val:
                    cell_prod.fill = COR_ERRO
                    mensagens.append("Produto vazio")
                elif len(prod_val) > 45:
                    cell_prod.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, 'Produto' excedeu o limite de caracteres")
                else:
                    cell_prod.fill = COR_VALIDO
                seen_produto[prod_val] = seen_produto.get(prod_val, 0) + 1

            # Validar CodFilial (usando índice direto)
            if idx_codfilial is None:
                mensagens.append("CodFilial ausente")
            else:
                cell_cf = row[idx_codfilial]
                cf_val = get_val(cell_cf)
                if cf_val and len(cf_val) > 40:
                    cell_cf.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, 'CodFilial' excedeu o limite de caracteres")
                if not cf_val and filial_unica:
                    cell_cf.value = filial_valor_unico
                    cf_val = filial_valor_unico
                    cell_cf.fill = COR_ADVERTENCIA
                    mensagens.append("Advertencia, CodFilial corrigido automaticamente")
                elif not cf_val:
                    cell_cf.fill = COR_ERRO
                    mensagens.append("CodFilial ausente e múltiplas opções disponíveis")
                elif cf_val not in filial_cod_list:
                    if filial_unica:
                        cell_cf.value = filial_valor_unico
                        cf_val = filial_valor_unico
                        cell_cf.fill = COR_ADVERTENCIA
                        mensagens.append("Advertencia, CodFilial corrigido automaticamente")
                    else:
                        cell_cf.fill = COR_ERRO
                        mensagens.append("CodFilial inexistente")
                else:
                    cell_cf.fill = COR_VALIDO

            # Validar CodFamilia (usando índice direto e Set)
            if idx_codfamilia is not None:
                cell_cfam = row[idx_codfamilia]
                cfam_val = get_val(cell_cfam)
                if cfam_val:
                    if not cfam_val.isdigit():
                        cell_cfam.fill = COR_ERRO
                        mensagens.append("CodFamilia deve ser inteiro")
                    elif cfam_val not in familia_cod_set:
                        cell_cfam.fill = COR_ERRO
                        mensagens.append("CodFamilia inexistente")
                    else:
                        cell_cfam.fill = COR_VALIDO
                else:
                    cell_cfam.fill = COR_VALIDO

            # Validar CodEstilo (usando índice direto e Set)
            if idx_codestilo is not None:
                cell_ce = row[idx_codestilo]
                ce_val = get_val(cell_ce)
                if ce_val:
                    if not ce_val.isdigit():
                        cell_ce.fill = COR_ERRO
                        mensagens.append("CodEstilo deve ser inteiro")
                    elif ce_val not in estilo_cod_set:
                        cell_ce.fill = COR_ERRO
                        mensagens.append("CodEstilo inexistente")
                    else:
                        cell_ce.fill = COR_VALIDO
                else:
                    cell_ce.fill = COR_VALIDO
            # Validar QtdeMultipla e QtdeMinima primeiro (regras simples)
            for campo in ["QtdeMultipla", "QtdeMinima"]:
                idx = header.get(campo)
                if idx is not None:
                    cell_q = row[idx]
                    if cell_q.value:
                        try:
                            q_val = int(cell_q.value)
                            if not (1 <= q_val <= 999999):
                                cell_q.fill = COR_ERRO
                                mensagens.append(f"{campo} fora do intervalo")
                            else:
                                cell_q.fill = COR_VALIDO
                        except:
                            cell_q.fill = COR_ERRO
                            mensagens.append(f"{campo} inválido")

            # Nova lógica para QtdeTabela1, QtdeTabela2, QtdeTabela3
            qtde1_idx = header.get("QtdeTabela1")
            qtde2_idx = header.get("QtdeTabela2")
            qtde3_idx = header.get("QtdeTabela3")
            
            qtde1_val = None
            qtde2_val = None
            qtde3_val = None
            qtde1_preenchida = False
            qtde2_preenchida = False
            qtde3_preenchida = False
            
            # Verificar quais QtdeTabela estão preenchidas
            if qtde1_idx is not None:
                cell_q1 = row[qtde1_idx]
                if cell_q1.value:
                    qtde1_preenchida = True
                    try:
                        qtde1_val = int(cell_q1.value)
                        if not (1 <= qtde1_val <= 999999):
                            cell_q1.fill = COR_ERRO
                            mensagens.append("QtdeTabela1 fora do intervalo")
                        else:
                            cell_q1.fill = COR_VALIDO
                    except:
                        cell_q1.fill = COR_ERRO
                        mensagens.append("QtdeTabela1 inválido")
                        qtde1_val = None
                        
            if qtde2_idx is not None:
                cell_q2 = row[qtde2_idx]
                if cell_q2.value:
                    qtde2_preenchida = True
                    try:
                        qtde2_val = int(cell_q2.value)
                        if not (1 <= qtde2_val <= 999999):
                            cell_q2.fill = COR_ERRO
                            mensagens.append("QtdeTabela2 fora do intervalo")
                        else:
                            cell_q2.fill = COR_VALIDO
                    except:
                        cell_q2.fill = COR_ERRO
                        mensagens.append("QtdeTabela2 inválido")
                        qtde2_val = None
                        
            if qtde3_idx is not None:
                cell_q3 = row[qtde3_idx]
                if cell_q3.value:
                    qtde3_preenchida = True
                    try:
                        qtde3_val = int(cell_q3.value)
                        if not (1 <= qtde3_val <= 999999):
                            cell_q3.fill = COR_ERRO
                            mensagens.append("QtdeTabela3 fora do intervalo")
                        else:
                            cell_q3.fill = COR_VALIDO
                    except:
                        cell_q3.fill = COR_ERRO
                        mensagens.append("QtdeTabela3 inválido")
                        qtde3_val = None

            # Validar padrão das QtdeTabela
            if qtde1_preenchida and not qtde2_preenchida:
                # QtdeTabela1 sozinha é inválido
                if qtde1_idx is not None:
                    row[qtde1_idx].fill = COR_ERRO
                mensagens.append("QtdeTabela1 não pode estar sozinha - QtdeTabela2 é obrigatória")
                
            elif qtde3_preenchida and not qtde2_preenchida:
                # QtdeTabela3 sem QtdeTabela2 é inválido
                if qtde3_idx is not None:
                    row[qtde3_idx].fill = COR_ERRO
                mensagens.append("QtdeTabela3 requer QtdeTabela1 e QtdeTabela2 preenchidas")
                
            elif qtde2_preenchida and not qtde1_preenchida:
                # QtdeTabela2 sem QtdeTabela1 é inválido
                if qtde2_idx is not None:
                    row[qtde2_idx].fill = COR_ERRO
                mensagens.append("QtdeTabela2 requer QtdeTabela1 preenchida")
                
            # Validar hierarquia das quantidades (se válidas)
            if qtde1_val and qtde2_val and qtde1_val >= qtde2_val:
                if qtde1_idx is not None:
                    row[qtde1_idx].fill = COR_ERRO
                if qtde2_idx is not None:
                    row[qtde2_idx].fill = COR_ERRO
                mensagens.append("QtdeTabela1 deve ser menor que QtdeTabela2")
                
            if qtde2_val and qtde3_val and qtde2_val >= qtde3_val:
                if qtde2_idx is not None:
                    row[qtde2_idx].fill = COR_ERRO
                if qtde3_idx is not None:
                    row[qtde3_idx].fill = COR_ERRO
                mensagens.append("QtdeTabela2 deve ser menor que QtdeTabela3")

            # Validar PrecoTabela1 (sempre obrigatório)
            cell_pt1 = self.get_mandatory_cell(row, header, "PrecoTabela1")
            pt1_val = None
            if cell_pt1 is None:
                mensagens.append("PrecoTabela1 ausente")
            else:
                try:
                    # Agora as fórmulas já foram convertidas para valores
                    if cell_pt1.value is None:
                        raise ValueError("Valor vazio")
                    elif isinstance(cell_pt1.value, (int, float)):
                        pt1_val = float(cell_pt1.value)
                        # Converter para formato com vírgula para manter consistência
                        cell_pt1.value = f"{pt1_val:.2f}".replace(".", ",")
                    else:
                        # Tratar como string e converter formato
                        valor_str = self.get_valor_string(cell_pt1)
                        
                        # Converter pontos para vírgulas inteligentemente
                        converted_value, foi_alterado = convert_price_to_comma_format(valor_str)
                        if foi_alterado:
                            mensagens.append(f"Advertencia: PrecoTabela1 corrigido de '{valor_str}' para '{converted_value}'")
                        
                        cell_pt1.value = converted_value  # Atualiza a célula com formato correto
                        pt1_val = float(converted_value.replace(",", "."))
                    
                    if not (0.01 <= pt1_val <= 999999.99):
                        cell_pt1.fill = COR_ERRO
                        mensagens.append("PrecoTabela1 fora do intervalo")
                    else:
                        cell_pt1.fill = COR_VALIDO
                except:
                    cell_pt1.fill = COR_ERRO
                    mensagens.append("PrecoTabela1 inválido")
                    pt1_val = None

            # Validar PrecoTabela2
            idx_pt2 = header.get("PrecoTabela2")
            pt2_val = None
            if idx_pt2 is not None:
                cell_pt2 = row[idx_pt2]
                if cell_pt2.value:
                    try:
                        # Agora as fórmulas já foram convertidas para valores
                        if cell_pt2.value is None:
                            raise ValueError("Valor vazio")
                        elif isinstance(cell_pt2.value, (int, float)):
                            pt2_val = float(cell_pt2.value)
                            # Converter para formato com vírgula para manter consistência
                            cell_pt2.value = f"{pt2_val:.2f}".replace(".", ",")
                        else:
                            # Tratar como string e converter formato
                            valor_str = self.get_valor_string(cell_pt2)
                            
                            # Converter pontos para vírgulas inteligentemente
                            converted_value, foi_alterado = convert_price_to_comma_format(valor_str)
                            if foi_alterado:
                                mensagens.append(f"Advertencia: PrecoTabela2 corrigido de '{valor_str}' para '{converted_value}'")
                            
                            cell_pt2.value = converted_value  # Atualiza a célula com formato correto
                            pt2_val = float(converted_value.replace(",", "."))
                        
                        if not (0.01 <= pt2_val <= 999999.99):
                            cell_pt2.fill = COR_ERRO
                            mensagens.append("PrecoTabela2 fora do intervalo")
                        else:
                            # Só aplicar hierarquia se QtdeTabela1 e QtdeTabela2 estiverem preenchidas
                            if (qtde1_preenchida and qtde2_preenchida) and pt1_val and pt2_val >= pt1_val:
                                cell_pt2.fill = COR_ERRO
                                mensagens.append("PrecoTabela2 deve ser menor que PrecoTabela1")
                            else:
                                cell_pt2.fill = COR_VALIDO
                    except:
                        cell_pt2.fill = COR_ERRO
                        mensagens.append("PrecoTabela2 inválido")
                        pt2_val = None

            # Validar PrecoTabela3
            idx_pt3 = header.get("PrecoTabela3")
            if idx_pt3 is not None:
                cell_pt3 = row[idx_pt3]
                if cell_pt3.value:
                    try:
                        # Agora as fórmulas já foram convertidas para valores
                        if cell_pt3.value is None:
                            raise ValueError("Valor vazio")
                        elif isinstance(cell_pt3.value, (int, float)):
                            pt3_val = float(cell_pt3.value)
                            # Converter para formato com vírgula para manter consistência
                            cell_pt3.value = f"{pt3_val:.2f}".replace(".", ",")
                        else:
                            # Tratar como string e converter formato
                            valor_str = self.get_valor_string(cell_pt3)
                            
                            # Converter pontos para vírgulas inteligentemente
                            converted_value, foi_alterado = convert_price_to_comma_format(valor_str)
                            if foi_alterado:
                                mensagens.append(f"Advertencia: PrecoTabela3 corrigido de '{valor_str}' para '{converted_value}'")
                            
                            cell_pt3.value = converted_value  # Atualiza a célula com formato correto
                            pt3_val = float(converted_value.replace(",", "."))
                        
                        if not (0.01 <= pt3_val <= 999999.99):
                            cell_pt3.fill = COR_ERRO
                            mensagens.append("PrecoTabela3 fora do intervalo")
                        else:
                            # Só aplicar hierarquia se QtdeTabela1, QtdeTabela2 e QtdeTabela3 estiverem preenchidas
                            if (qtde1_preenchida and qtde2_preenchida and qtde3_preenchida) and pt2_val and pt3_val >= pt2_val:
                                cell_pt3.fill = COR_ERRO
                                mensagens.append("PrecoTabela3 deve ser menor que PrecoTabela2")
                            else:
                                cell_pt3.fill = COR_VALIDO
                    except:
                        cell_pt3.fill = COR_ERRO
                        mensagens.append("PrecoTabela3 inválido")
            idx = header.get("LimiteDescIndividual")
            if idx is not None:
                cell_lim = row[idx]
                if cell_lim.value:
                    try:
                        lim_str = self.get_valor_string(cell_lim)
                        lim_val = float(lim_str.replace(",", "."))
                        if not (0.00 <= lim_val <= 99.99):
                            cell_lim.fill = COR_ERRO
                            mensagens.append("LimiteDescIndividual fora do intervalo")
                        else:
                            cell_lim.fill = COR_VALIDO
                            cell_lim.value = f"{lim_val:.2f}".replace(".", ",")
                    except:
                        cell_lim.fill = COR_ERRO
                        mensagens.append("LimiteDescIndividual inválido")
            multiplo_grade = None
            idx = header.get("MultiploGrade")
            if idx is not None:
                cell_mg = row[idx]
                if cell_mg.value:
                    try:
                        multiplo_grade = int(cell_mg.value)
                        if not (1 <= multiplo_grade <= 999999):
                            cell_mg.fill = COR_ERRO
                            mensagens.append("MultiploGrade fora do intervalo")
                        else:
                            cell_mg.fill = COR_VALIDO
                    except:
                        cell_mg.fill = COR_ERRO
                        mensagens.append("MultiploGrade inválido")
            if multiplo_grade is not None:
                idx = header.get("DescontoGrade")
                if idx is not None:
                    cell_dg = row[idx]
                    if cell_dg.value:
                        try:
                            dg_str = self.get_valor_string(cell_dg)
                            dg_val = float(dg_str.replace(",", "."))
                            if not (0.00 <= dg_val <= 99.99):
                                cell_dg.fill = COR_ERRO
                                mensagens.append("DescontoGrade fora do intervalo")
                            else:
                                cell_dg.fill = COR_VALIDO
                                cell_dg.value = f"{dg_val:.2f}".replace(".", ",")
                        except:
                            cell_dg.fill = COR_ERRO
                            mensagens.append("DescontoGrade inválido")
            idx = header.get("PrecoPromocional")
            if idx is not None:
                cell_pp = row[idx]
                if corrigir_campo(cell_pp, {"S", "s", "N", "n", ""}):
                    cell_pp.fill = COR_ERRO
                    mensagens.append("PrecoPromocional inválido")
                else:
                    cell_pp.fill = COR_VALIDO
            idx = header.get("AliquotaIPI")
            if idx is not None:
                cell_ipi = row[idx]
                if cell_ipi.value:
                    try:
                        # Converter para string primeiro, depois tratar
                        ipi_str = self.get_valor_string(cell_ipi)
                        ipi_val = float(ipi_str.replace(",", "."))
                        if not (0.00 <= ipi_val <= 99.99):
                            cell_ipi.fill = COR_ERRO
                            mensagens.append("AliquotaIPI fora do intervalo")
                        else:
                            cell_ipi.fill = COR_VALIDO
                            cell_ipi.value = f"{ipi_val:.2f}".replace(".", ",")
                    except:
                        cell_ipi.fill = COR_ERRO
                        mensagens.append("AliquotaIPI inválido")
            idx = header.get("TipoVendaSemEstoque")
            if idx is not None:
                cell_tv = row[idx]
                if corrigir_campo(cell_tv, {"L", "l", "b", "B", "c", "C", ""}):
                    cell_tv.fill = COR_ERRO
                    mensagens.append("TipoVendaSemEstoque inválido")
                else:
                    cell_tv.fill = COR_VALIDO
            idx = header.get("QtdeEstoqueAtual")
            if idx is not None:
                cell_qea = row[idx]
                if cell_qea.value:
                    try:
                        qea_val = int(cell_qea.value)
                        if not (1 <= qea_val <= 999999):
                            cell_qea.fill = COR_ERRO
                            mensagens.append("QtdeEstoqueAtual fora do intervalo")
                        else:
                            cell_qea.fill = COR_VALIDO
                    except:
                        cell_qea.fill = COR_ERRO
                        mensagens.append("QtdeEstoqueAtual inválido")
            idx = header.get("QtdeEstoqueFuturo")
            if idx is not None:
                cell_qef = row[idx]
                if cell_qef.value:
                    try:
                        qef_val = int(cell_qef.value)
                        if qef_val == 0:
                            cell_qef.value = ""
                            cell_qef.fill = COR_VALIDO
                        elif not (1 <= qef_val <= 999999):
                            cell_qef.fill = COR_ERRO
                            mensagens.append("QtdeEstoqueFuturo fora do intervalo")
                        else:
                            cell_qef.fill = COR_VALIDO
                    except:
                        cell_qef.fill = COR_ERRO
                        mensagens.append("QtdeEstoqueFuturo inválido")
            idx = header.get("DtEstoqueFuturo")
            if idx is not None:
                cell_data = row[idx]
                if cell_data.value:
                    try:
                        dt = None
                        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                            try:
                                dt = datetime.strptime(self.get_valor_string(cell_data), fmt)
                                break
                            except:
                                continue
                        if dt:
                            cell_data.value = dt
                            cell_data.number_format = "DD/MM/YYYY"
                            cell_data.fill = COR_VALIDO
                        else:
                            cell_data.fill = COR_ERRO
                            mensagens.append("DtEstoqueFuturo com formato inválido")
                    except:
                        cell_data.fill = COR_ERRO
                        mensagens.append("DtEstoqueFuturo inválida")
                    idx_qef = header.get("QtdeEstoqueFuturo")
                    if idx_qef is not None:
                        cell_qef = row[idx_qef]
                        if not cell_qef.value or str(cell_qef.value).strip() == "":
                            mensagens.append(
                                "Advertência(s): QtdeEstoqueFuturo não contém saldo suficiente para a DtEstoqueFuturo"
                            )
            idx = header.get("PathFotografia")
            if idx is not None:
                cell_pf = row[idx]
                if cell_pf.value:
                    path_foto = self.get_valor_string(cell_pf)
                    
                    # Validação de tamanho (existente)
                    if len(path_foto) > 60:
                        cell_pf.fill = COR_ADVERTENCIA
                        mensagens.append("Advertencia, 'PathFotografia' excedeu o limite de caracteres")
                    
                    # Validação EXTREMAMENTE RÁPIDA de existência de arquivo
                    elif self.cache_arquivos_imagem is not None and path_foto not in self.cache_arquivos_imagem:
                        cell_pf.fill = COR_ADVERTENCIA
                        mensagens.append(f"Advertencia: '{path_foto}' não existe na pasta C:\\Users\\Public\\Documents\\SRPP\\Imagens_Produto")
                    
                    else:
                        cell_pf.fill = COR_VALIDO
                else:
                    cell_pf.fill = COR_VALIDO
            idx = header.get("QtdeEtiquetas")
            if idx is not None:
                cell_qe = row[idx]
                if cell_qe.value:
                    try:
                        qe_val = int(cell_qe.value)
                        if not (1 <= qe_val <= 999):
                            cell_qe.fill = COR_ERRO
                            mensagens.append("QtdeEtiquetas fora do intervalo")
                        else:
                            cell_qe.fill = COR_VALIDO
                    except:
                        cell_qe.fill = COR_ERRO
                        mensagens.append("QtdeEtiquetas inválido")
            resultado_fill = self.determinar_fill_resultado(mensagens)
            for cell in row:
                cell.fill = resultado_fill
            dup_col = header.get("Duplicados")
            if dup_col is not None:
                dup_cell = row[dup_col]
                if dup_valores:
                    dup_cell.value = ";".join(dup_valores)
                    dup_cell.fill = COR_DUPLICADO
                else:
                    dup_cell.value = ""
            linhas_erros += 1 if resultado_fill == COR_ERRO else 0
            linhas_validas += 1 if resultado_fill == COR_VALIDO else 0
            linhas_advertencias += 1 if resultado_fill == COR_ADVERTENCIA else 0


            mensagem_resultado = "; ".join(mensagens)
            if not mensagem_resultado.strip():
                mensagem_resultado = "Validado com sucesso!"
            self.escrever_resultado_linha(
                sheet, row[0].row, mensagem_resultado, result_col, resultado_fill
            )
            # Aplicar negrito caso seja "Validado com sucesso!"
            if mensagem_resultado == "Validado com sucesso!":
                cell = sheet.cell(row=row[0].row, column=result_col)
                cell.font = Font(bold=True)
                
        duplicados_vazios = True
        if "Duplicados" in header:
            for row in sheet.iter_rows(
                min_row=2,
                min_col=header.get("Duplicados") + 1,
                max_col=header.get("Duplicados") + 1,
            ):
                for cell in row:
                    if cell.value and str(cell.value).strip() != "":
                        duplicados_vazios = False
                        break
            if duplicados_vazios:
                sheet.delete_cols(header.get("Duplicados") + 1)
        self.gerar_status_por_aba(
            "PRODUTOS", total_linhas, linhas_validas, linhas_advertencias, linhas_erros
        )
        self.aplicar_borda(sheet)
        max_length = 0
        for row in sheet.iter_rows(min_row=2, min_col=result_col, max_col=result_col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(result_col)].width = max_length * 1.2
        return None


    def obter_nome_empresa(self):
        """Obtém o nome da empresa da célula C5 da aba EMPRESA. Se não existir, retorna 'erro'."""
        try:
            sheet = self.wb["EMPRESA"]
            nome_empresa = sheet["C5"].value
            return nome_empresa.strip() if nome_empresa else "erro"
        except KeyError:
            return "erro"




    def gerar_planilha_etiquetas(self):
        """
        Gera a planilha de etiquetas a partir da aba 'PRODUTOS', mas somente se houver
        pelo menos uma linha com QtdeEtiquetas > 0.
        A nova planilha conterá somente as seguintes colunas:
        CodProduto, CodAuxiliarProduto, Produto - Linha 1, Produto - Linha 2,
        QtdeMultipla, QtdeMinima, QtdeTabela1, QtdeTabela2, QtdeTabela3,
        PrecoTabela1, PrecoTabela2, PrecoTabela3, AliquotaIPI, QtdeEtiquetas.
        A coluna 'Produto' original é substituída por 'Produto - Linha 1' e 'Produto - Linha 2',
        dividindo o texto sem cortar palavras (máximo 23 caracteres por linha).
        Todas as células são formatadas como texto, com fonte Arial 10; o cabeçalho é negrito,
        centralizado e com fundo amarelo (#FFFF00); os dados são alinhados à esquerda.
        Retorna o caminho do arquivo gerado ou None se não houver etiquetas.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            import os
            from datetime import datetime

            # Tenta obter a aba 'PRODUTOS'
            try:
                produtos_sheet = self.wb["PRODUTOS"]
            except KeyError:
                return None

            # Define o cabeçalho fixo para a planilha de etiquetas
            final_header = [
                "CodProduto",
                "CodAuxiliarProduto",
                "Produto - Linha 1",
                "Produto - Linha 2",
                "QtdeMultipla",
                "QtdeMinima",
                "QtdeTabela1",
                "QtdeTabela2",
                "QtdeTabela3",
                "PrecoTabela1",
                "PrecoTabela2",
                "PrecoTabela3",
                "AliquotaIPI",
                "QtdeEtiquetas"
            ]

            # Obter o cabeçalho original da aba PRODUTOS (primeira linha)
            header = [cell.value for cell in produtos_sheet[1]]
            # Campos obrigatórios na planilha original
            required_fields = [
                "CodProduto",
                "CodAuxiliarProduto",
                "Produto",
                "QtdeMultipla",
                "QtdeMinima",
                "QtdeTabela1",
                "QtdeTabela2",
                "QtdeTabela3",
                "PrecoTabela1",
                "PrecoTabela2",
                "PrecoTabela3",
                "AliquotaIPI",
                "QtdeEtiquetas"
            ]
            header_map = {col: idx for idx, col in enumerate(header)}
            for field in required_fields:
                if field not in header_map:
                    return None

            idx_qtde = header_map["QtdeEtiquetas"]
            idx_produto = header_map["Produto"]

            # Filtrar linhas com QtdeEtiquetas > 0
            linhas_etiquetas = []
            for row in produtos_sheet.iter_rows(min_row=2, values_only=True):
                qtde = row[idx_qtde]
                if qtde is None or str(qtde).strip() == "":
                    continue
                try:
                    qtde_val = float(qtde)
                except ValueError:
                    continue
                if qtde_val > 0:
                    linhas_etiquetas.append(row)

            if not linhas_etiquetas:
                return None

            wb_etiquetas = Workbook()
            sheet_etiquetas = wb_etiquetas.active
            sheet_etiquetas.title = "Etiquetas"

            # Adiciona o cabeçalho fixo
            sheet_etiquetas.append(final_header)

            # Processa cada linha filtrada e monta a nova linha com as colunas adequadas
            for row in linhas_etiquetas:
                nova_linha = []
                for col in final_header:
                    if col in ["Produto - Linha 1", "Produto - Linha 2"]:
                        # Obtém o valor original da coluna 'Produto'
                        produto_val = row[idx_produto]
                        produto_str = str(produto_val) if produto_val is not None else ""
                        # Usa a função split_text para dividir o texto (verifica se self.split_text existe)
                        if hasattr(self, "split_text"):
                            linha1, linha2 = self.split_text(produto_str, 23)
                        else:
                            linha1, linha2 = split_text(produto_str, 23)
                        nova_linha.append(linha1 if col == "Produto - Linha 1" else linha2)
                    else:
                        # Copia o valor da coluna correspondente, se existir
                        if col in header_map:
                            nova_linha.append(row[header_map[col]])
                        else:
                            nova_linha.append("")
                sheet_etiquetas.append(nova_linha)

            # Formatação: Cabeçalho: Arial 10, negrito, centralizado, fundo amarelo (#FFFF00);
            # dados: Arial 10, alinhados à esquerda
            from openpyxl.utils import get_column_letter

            font_header = Font(name="Arial", size=10, bold=True)
            font_data = Font(name="Arial", size=10, bold=False)
            align_header = Alignment(horizontal="center", vertical="center")
            align_data = Alignment(horizontal="left", vertical="center")
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Formatar cabeçalho (linha 1)
            for cell in sheet_etiquetas[1]:
                cell.font = font_header
                cell.alignment = align_header
                cell.fill = header_fill
                cell.number_format = "@"  # formato texto

            # Formatar os dados (a partir da linha 2)
            for row in sheet_etiquetas.iter_rows(min_row=2):
                for cell in row:
                    cell.font = font_data
                    cell.alignment = align_data
                    cell.number_format = "@"  # forçar como texto

            # Ajustar a largura de cada coluna com base no maior conteúdo
            for col in sheet_etiquetas.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        try:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                        except Exception:
                            pass
                sheet_etiquetas.column_dimensions[column].width = max_length + 2

            # Define o nome do arquivo de etiquetas usando self.emp_nome e timestamp
            nome_base = self.emp_nome if self.emp_nome and self.emp_nome.strip() else "erro"
            timestamp = datetime.now().strftime("%Y.%m.%d %H-%M")  # novo formato: Ano.Mês.Dia Hora-Minuto
            nome_arquivo = f"{timestamp}_{nome_base}_ETIQUETAS.xls"
            
            # Salva o workbook em um buffer de memória em vez de um arquivo físico
            from io import BytesIO
            etiquetas_data = BytesIO()
            wb_etiquetas.save(etiquetas_data)
            etiquetas_data.seek(0)  # Retorna o ponteiro para o início do buffer
            
            # Retorna o objeto BytesIO e o nome do arquivo
            return etiquetas_data, nome_arquivo


        except Exception as e:
            print(f"Erro ao gerar planilha de etiquetas: {e}")
            return None

    
    #def salvar_planilha(self, novo_arquivo=None):
            # Alterado para salvar na pasta 'build' em vez da área de trabalho
     #       self.wb.active = self.wb["RESULTADO DAS VALIDAÇÕES"]

            # Obtém o nome base da empresa ou usa um padrão
     #       nome_base = (self.emp_nome or "Planilha_Validada").strip()

            # Gera timestamp no formato desejado: Ano.Mês.Dia Hora-Minuto
      #      timestamp = datetime.now().strftime("%Y.%m.%d %H-%M")

            # Se não for passado um nome de arquivo, cria um nome padrão conforme o novo formato
       #     if novo_arquivo is None:
                # O nome do arquivo ficará: timestamp + _ + nome_base + _IMPORTAÇÃO.xlsx
        #        novo_arquivo = f"{timestamp}_{nome_base}_IMPORTAÇÃO.xlsx"

            # Define o diretório para salvar a planilha (pasta 'build')
         #   diretorio_build = os.path.join(os.getcwd(), "build")
          #  if not os.path.exists(diretorio_build):
           #     os.makedirs(diretorio_build)
            #caminho_arquivo = os.path.join(diretorio_build, novo_arquivo)

            # Salva a planilha no caminho especificado
            #self.wb.save(caminho_arquivo)

            # Retorna o caminho do arquivo salvo
            #return caminho_arquivo
        
    def _timing(self, nome, funcao, *args, **kwargs):
        """Executa função e registra tempo se em modo dev."""
        import time
        if getattr(self, '_dev_mode', False):
            t0 = time.perf_counter()
            result = funcao(*args, **kwargs)
            self._timings[nome] = time.perf_counter() - t0
            return result
        return funcao(*args, **kwargs)

    def processar(self, empresa):
        """
        Processa a validação e retorna os dados em memória.
        Retorna: (dados_excel, nome_arquivo, status, resultados)
        """
        # Lista de etapas de validação com seus nomes amigáveis
        # PRODUTOS é tratado separadamente pois tem progresso granular
        etapas_pre_produtos = [
            (self.limpar_planilha, "limpar_planilha", "Limpando planilha..."),
            (self.validar_EMPRESA, "validar_EMPRESA", "Validando EMPRESA..."),
            (self.pre_validar_filial, "pre_validar_filial", "Pré-validando FILIAL..."),
            (self.validar_FILIAL, "validar_FILIAL", "Validando FILIAL..."),
            (self.validar_REPR, "validar_REPR", "Validando REPRESENTANTES..."),
            (self.validar_PAGTO, "validar_PAGTO", "Validando PAGAMENTO..."),
            (self.validar_PAGTOFILIAL, "validar_PAGTOFILIAL", "Validando PAGTO x FILIAL..."),
            (self.validar_TRANSP, "validar_TRANSP", "Validando TRANSPORTADORAS..."),
            (self.validar_ESTADOS, "validar_ESTADOS", "Validando ESTADOS..."),
            (self.validar_CLIENTES, "validar_CLIENTES", "Validando CLIENTES..."),
            (self.validar_FAMILIAS, "validar_FAMILIAS", "Validando FAMÍLIAS..."),
            (self.validar_ESTILOS, "validar_ESTILOS", "Validando ESTILOS..."),
        ]

        # Etapas pré-PRODUTOS: 5% a 50%
        total_pre = len(etapas_pre_produtos)
        for i, (funcao, nome, mensagem) in enumerate(etapas_pre_produtos):
            percentual = 5 + int((i / total_pre) * 45)
            self._reportar_progresso(percentual, mensagem)
            self._timing(nome, funcao)

        # PRODUTOS: 50% a 88% (range de 38% para progresso granular)
        self._reportar_progresso(50, "Validando PRODUTOS...")
        self._timing("validar_PRODUTOS", self.validar_PRODUTOS, progress_base=50, progress_range=38)

        # Relatório final: 88% a 90%
        self._reportar_progresso(88, "Gerando relatório final...")
        self._timing("gerar_relatorio_final", self.gerar_relatorio_final)

        self._reportar_progresso(92, "Determinando status...")

        # Determinar o status com base em erros e advertências
        if any(dados["erros"] > 0 for dados in self.resultados_validacao.values()):
            status = "reprovado"
        elif any(dados["advertencias"] > 0 for dados in self.resultados_validacao.values()):
            status = "advertencias"
        else:
            status = "aprovado"

        self._reportar_progresso(95, "Salvando arquivo...")

        # Prepara o nome do arquivo
        timestamp = datetime.now().strftime("%Y.%m.%d %H-%M")
        nome_arquivo = f"{timestamp}_{self.emp_nome}_IMPORTAÇÃO.xlsx"

        # Salva o workbook em um buffer de memória
        from io import BytesIO
        excel_data = BytesIO()
        self.wb.save(excel_data)
        excel_data.seek(0)  # Retorna o ponteiro para o início do buffer

        resultados = [
            {"Planilha": aba, **dados}
            for aba, dados in self.resultados_validacao.items()
        ]

        self._reportar_progresso(100, "Concluído!")

        return excel_data, nome_arquivo, status, resultados

