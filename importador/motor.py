"""
Motor de importacao - executa as stored procedures no SQL Server.
"""

import os
import time
from datetime import datetime
import pyodbc
from openpyxl import load_workbook
from mapeamento import MAPA_ABAS, ORDEM_IMPORTACAO
from sql_batch import (
    SQL_REMOVE_SYNC_TRIGGERS,
    SQL_CREATE_STAGING,
    SQL_CREATE_PROCEDURE,
    SQL_CREATE_PROCEDURE_BODY,
    SQL_INSERT_STAGING,
    # Transportadora
    SQL_CREATE_STAGING_TRANSPORTADORA,
    SQL_CREATE_PROCEDURE_TRANSPORTADORA,
    SQL_CREATE_PROCEDURE_BODY_TRANSPORTADORA,
    SQL_INSERT_STAGING_TRANSPORTADORA,
    # Cliente
    SQL_CREATE_STAGING_CLIENTE,
    SQL_CREATE_PROCEDURE_CLIENTE,
    SQL_CREATE_PROCEDURE_BODY_CLIENTE,
    SQL_INSERT_STAGING_CLIENTE,
)


class PlanilhaImportador:

    # Drivers que suportam fast_executemany sem problemas
    DRIVERS_FAST_EXECUTEMANY = [
        "ODBC Driver 18",
        "ODBC Driver 17",
        "ODBC Driver 13",
        "ODBC Driver 11",
    ]

    def __init__(self, connection_string, progress_callback=None, log_callback=None):
        self.connection_string = connection_string
        self.progress_callback = progress_callback
        self.log_callback = log_callback
        self.conn = None
        self.cancelar = False
        self._fast_executemany_ok = self._verificar_fast_executemany()

    def _verificar_fast_executemany(self):
        """Verifica se o driver suporta fast_executemany."""
        for driver in self.DRIVERS_FAST_EXECUTEMANY:
            if driver in self.connection_string:
                return True
        return False

    def _log(self, msg):
        if self.log_callback:
            self.log_callback(msg)
        print(msg)

    def _progresso(self, percentual, mensagem):
        if self.progress_callback:
            self.progress_callback(percentual, mensagem)

    def _normalizar_nomes_abas(self, wb):
        """Corrige a caixa (maiuscula/minuscula) do nome das abas para o padrao
        esperado (ex: 'Produtos' -> 'PRODUTOS'), pois o resto do codigo busca
        as abas pelo nome exato em maiusculas."""
        canonicos_por_upper = {nome.upper(): nome for nome in ORDEM_IMPORTACAO}
        for sheet_name in list(wb.sheetnames):
            nome_canonico = canonicos_por_upper.get(sheet_name.strip().upper())
            if nome_canonico and sheet_name != nome_canonico and nome_canonico not in wb.sheetnames:
                ws = wb[sheet_name]
                # Renomeia em dois passos: o openpyxl trata nomes de aba como
                # case-insensitive unicos e recusa "Produtos" -> "PRODUTOS"
                # direto (vira "PRODUTOS1"), pois ve como duplicata de si mesma.
                ws.title = f"~~tmp~~{nome_canonico}"
                ws.title = nome_canonico

    # ----------------------------------------------------------
    # Conexao
    # ----------------------------------------------------------

    def conectar(self):
        self.conn = pyodbc.connect(self.connection_string, autocommit=True)
        cursor = self.conn.cursor()
        cursor.execute("SELECT DB_NAME()")
        self.database_name = cursor.fetchone()[0]
        cursor.close()


    def desconectar(self):
        if self.conn:
            try:
                self.conn.close()
            except Exception:
                pass
            self.conn = None

    @staticmethod
    def testar_conexao(connection_string):
        """Testa conexao e retorna (ok, mensagem)."""
        try:
            conn = pyodbc.connect(connection_string, autocommit=True, timeout=10)
            cursor = conn.cursor()
            cursor.execute("SELECT @@SERVERNAME")
            nome = cursor.fetchone()[0]
            conn.close()
            return True, f"Conectado a: {nome}"
        except Exception as e:
            return False, str(e)

    # ----------------------------------------------------------
    # Helpers SQL
    # ----------------------------------------------------------

    def _converter_valor(self, valor, tipo_sql):
        """Converte valor da celula Excel para tipo SQL.
        Retorna tipos Python consistentes para evitar erro 22018 no fast_executemany:
        - int/smallint/tinyint -> int ou None
        - decimal -> float ou None
        - datetime -> datetime object ou None (nao string!)
        - varchar/char -> str ou None
        """
        if valor is None:
            return None

        # datetime: preservar objeto datetime do openpyxl (nao converter pra string!)
        if tipo_sql == "datetime":
            from datetime import datetime as dt
            if isinstance(valor, dt):
                return valor
            # Se for string, tentar parsear
            val_str = str(valor).strip()
            if val_str == "":
                return None
            try:
                # Tentar formatos comuns
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
                    try:
                        return dt.strptime(val_str, fmt)
                    except ValueError:
                        continue
                return None  # Formato desconhecido -> NULL
            except Exception:
                return None

        val_str = str(valor).strip()
        if val_str == "":
            return None
        try:
            if tipo_sql in ("int", "smallint", "tinyint"):
                val = int(float(val_str.replace(",", ".")))
                if tipo_sql == "tinyint" and not (0 <= val <= 255):
                    return None
                return val
            elif tipo_sql == "decimal":
                return float(val_str.replace(",", "."))
            else:
                return val_str.upper()
        except (ValueError, TypeError):
            # Tipo numerico que nao pode ser convertido -> NULL
            # (evita erro 22018 no bulk insert para staging)
            if tipo_sql in ("int", "smallint", "tinyint", "decimal"):
                return None
            return val_str

    def _formatar_cnpjcpf(self, valor):
        """Formata CNPJ para formato SRPP (NNN.NNN.NNN/NNNN-NN, 19 chars).
        CPF (11 digitos) e convertido para formato CNPJ com zeros a esquerda.
        Retorna NULL se nao houver digitos suficientes (minimo 11 para CPF).
        """
        if valor is None:
            return None
        val = str(valor).strip()
        if not val:
            return None
        # Extrair digitos para validacao
        digitos = ''.join(c for c in val if c.isdigit())
        # Se nao tem pelo menos 11 digitos, retorna NULL (invalido)
        if len(digitos) < 11:
            return None
        # Ja formatado 19 chars (formato SRPP)
        if len(val) == 19 and '/' in val:
            return val
        # Formatado padrao brasileiro 18 chars -> pad para 19
        if len(val) == 18 and '/' in val:
            return '0' + val
        # CNPJ 14 digitos -> formatar para SRPP
        if len(digitos) == 14:
            d = digitos.zfill(15)
            return f"{d[0:3]}.{d[3:6]}.{d[6:9]}/{d[9:13]}-{d[13:15]}"
        # CPF (11 digitos) -> pad para 14 e formatar como CNPJ SRPP
        if len(digitos) == 11:
            digitos = digitos.zfill(14)
            d = digitos.zfill(15)
            return f"{d[0:3]}.{d[3:6]}.{d[6:9]}/{d[9:13]}-{d[13:15]}"
        # 12 ou 13 digitos - pad para 14 e formatar
        if len(digitos) in (12, 13):
            digitos = digitos.zfill(14)
            d = digitos.zfill(15)
            return f"{d[0:3]}.{d[3:6]}.{d[6:9]}/{d[9:13]}-{d[13:15]}"
        # Mais de 14 digitos - pegar os ultimos 14 e formatar
        if len(digitos) > 14:
            digitos = digitos[-14:]
            d = digitos.zfill(15)
            return f"{d[0:3]}.{d[3:6]}.{d[6:9]}/{d[9:13]}-{d[13:15]}"
        # Fallback: retorna NULL para evitar erro de constraint
        return None

    # Campos onde valor 0 deve ser convertido para NULL (constraint > 0 OR NULL)
    _ZERO_TO_NULL = {"@ddd", "@codtransportadora", "@codrepresentante", "@qtdetabela1", "@qtdetabela2", "@qtdetabela3", "@codfamilia", "@codestilo", "@qtdemultipla", "@qtdeminima", "@qtdeetiquetas", "@precotabela", "@desconto1", "@desconto2", "@desconto3", "@vlrminimopedido", "@precotabela2", "@precotabela3", "@qtdeestoqueatual", "@qtdeestoquefuturo", "@limitedescindividual", "@multiplograde", "@descontograde"}

    # ==================== CONSTRAINTS CHECK DO BANCO ====================
    # Mapeamento das constraints CHECK para validacao pre-batch
    # Tipos de constraint:
    #   "maior_que_zero": valor > 0 (obrigatorio)
    #   "maior_que_zero_ou_null": valor > 0 OU NULL (corrigivel -> NULL)
    #   "maior_igual_zero": valor >= 0 (obrigatorio)
    #   "maior_igual_zero_ou_null": valor >= 0 OU NULL
    #   "nao_vazio": string nao vazia (obrigatorio)
    #   "nao_vazio_ou_null": string nao vazia OU NULL (corrigivel -> NULL)
    #   "formato_cnpjcpf": formato NNN.NNN.NNN/NNNN-NN ou NULL (corrigivel -> NULL)
    #   "formato_cep": formato NNNNN-NNN ou NULL (corrigivel -> NULL)
    #   "uf_valida": sigla de estado valida ou NULL
    #   "sim_nao": 'S' ou 'N' ou NULL

    _UFS_VALIDAS = {'AC', 'AL', 'AM', 'AP', 'BA', 'CE', 'DF', 'ES', 'EX', 'GO', 'MA', 'MG', 'MS', 'MT', 'PA', 'PB', 'PE', 'PI', 'PR', 'RJ', 'RN', 'RO', 'RR', 'RS', 'SC', 'SE', 'SP', 'TO'}

    _CONSTRAINTS_TRANSPORTADORA = {
        "CodTransportadora": "maior_que_zero",
        "Transportadora": "nao_vazio",
    }

    _CONSTRAINTS_CLIENTE = {
        "CodCliente": "maior_que_zero",
        "NomeFantasia": "nao_vazio",
        "CNPJCPF": "formato_cnpjcpf",
        "CEP": "formato_cep",
        "CodRepresentante": "maior_que_zero_ou_null",
        "DDD": "maior_que_zero_ou_null",
        "UF": "uf_valida",
        "Bairro": "nao_vazio_ou_null",
        "Cidade": "nao_vazio_ou_null",
        "EMail": "nao_vazio_ou_null",
        "FAX": "nao_vazio_ou_null",
        "IERG": "nao_vazio_ou_null",
        "Logradouro": "nao_vazio_ou_null",
        "NomeContato": "nao_vazio_ou_null",
        "Observacao": "nao_vazio_ou_null",
        "RazaoSocial": "nao_vazio_ou_null",
        "Telefone1": "nao_vazio_ou_null",
        "Telefone2": "nao_vazio_ou_null",
    }

    _CONSTRAINTS_PRODUTO = {
        "CodProduto": "nao_vazio",
        "Produto": "nao_vazio",
        "PrecoTabela1": "maior_igual_zero",
        "CodAuxiliarProduto": "nao_vazio_ou_null",
        "PathFotografia": "nao_vazio_ou_null",
        "PrecoTabela2": "maior_igual_zero_ou_null",
        "PrecoTabela3": "maior_igual_zero_ou_null",
        "AliquotaIPI": "maior_igual_zero_ou_null",
        "DescontoGrade": "maior_igual_zero_ou_null",
        "LimiteDescIndividual": "maior_igual_zero_ou_null",
        "MultiploGrade": "maior_igual_zero_ou_null",
        "QtdeEstoqueAtual": "maior_igual_zero_ou_null",
        "QtdeEstoqueFuturo": "maior_igual_zero_ou_null",
        "QtdeMinima": "maior_igual_zero_ou_null",
        "QtdeMultipla": "maior_igual_zero_ou_null",
        "QtdeTabela1": "maior_igual_zero_ou_null",
        "QtdeTabela2": "maior_igual_zero_ou_null",
        "QtdeTabela3": "maior_igual_zero_ou_null",
    }

    # Tamanhos maximos de campos varchar/char (baseado na estrutura REAL do banco SRPP)
    # Formato: "NomeColunaExcel": tamanho_maximo
    _MAX_LENGTHS = {
    # FILIAL
    "Filial": 40,
    "TituloAdicional1": 70,
    "TituloAdicional2": 70,
    "Logotipo": 50,

    # REPRESENTANTE
    "Representante": 20,

    # CONDIÇÃO DE PAGAMENTO
    "CondPagamento": 20,
    "TipoCondPagamento": 1,
    "CondPagamentoPadrao": 1,
    "VlrMinimoComEstAtual": 1,
    "VlrMinimoComEstFuturo": 1,
    "VlrMinimoComEstEsgotado": 1,

    # CONDIÇÃO DE PAGAMENTO FILIAL
    # (se futuramente tiver campos específicos, entram aqui)

    # TRANSPORTADORA
    "Transportadora": 20,
    "TransportadoraPadrao": 1,

    # ESTADOS
    "SiglaEstado": 2,
    "NomeEstado": 20,

    # CLIENTES
    "NomeFantasia": 20,
    "RazaoSocial": 40,
    
    "IERG": 19,
    "Logradouro": 40,
    "Bairro": 20,
    "Cidade": 20,
    "UF": 2,
    "CEP": 9,
    "Telefone1": 9,
    "Telefone2": 9,
    "FAX": 9,
    "NomeContato": 40,
    "NomeTransportadora": 20,
    "Observacao": 20,
    "EMail": 40,

    # FAMÍLIA
    "Familia": 45,

    # ESTILOS
    "Estilo": 45,

    # PRODUTOS
    "CodProduto": 20,
    "CodAuxiliarProduto": 20,
    "Produto": 40,
    "PathFotografia": 60,
    "PrecoPromocional": 1,
    "TipoVendaSemEstoque": 1,
}

    # ==================== FUNCOES DE PRE-VALIDACAO ====================

    def _validar_cnpjcpf_formato(self, valor):
        """
        Valida se CNPJCPF pode ser convertido para formato SRPP valido.
        Retorna (valido, valor_corrigido_ou_none, mensagem_erro).
        """
        if valor is None:
            return True, None, None
        val = str(valor).strip()
        if not val:
            return True, None, None

        # Extrair apenas digitos
        digitos = ''.join(c for c in val if c.isdigit())

        # Precisa de pelo menos 11 digitos (CPF) para ser valido
        if len(digitos) < 11:
            return False, None, f"CNPJCPF '{val}' invalido (apenas {len(digitos)} digitos, minimo 11)"

        return True, valor, None

    def _validar_cep_formato(self, valor):
        """
        Valida se CEP pode ser convertido para formato SRPP valido.
        Aceita CEPs com 5-8 digitos (faz pad com zeros a esquerda).
        Retorna (valido, valor_corrigido_ou_none, mensagem_erro).
        """
        if valor is None:
            return True, None, None
        val = str(valor).strip()
        if not val:
            return True, None, None

        digitos = ''.join(c for c in val if c.isdigit())
        # CEP precisa ter pelo menos 5 digitos para ser valido (ex: 01310 -> 01310-000)
        # Se tiver apenas 1-4 digitos ou for "0", provavelmente nao e CEP valido
        if len(digitos) < 5 or (len(digitos) == 1 and digitos == "0"):
            return False, None, f"CEP '{val}' invalido (apenas {len(digitos)} digitos, minimo 5)"

        return True, valor, None

    def _validar_constraint(self, valor, tipo_constraint, col_name):
        """
        Valida um valor contra um tipo de constraint.
        Retorna (valido, valor_corrigido, mensagem_erro, corrigivel).
        - valido: True se passou na validacao
        - valor_corrigido: valor ajustado (ou None se invalido e corrigivel)
        - mensagem_erro: descricao do problema se invalido
        - corrigivel: True se o valor pode ser automaticamente corrigido para NULL
        """
        # Normalizar valor
        if isinstance(valor, str):
            valor = valor.strip()
            if valor == "":
                valor = None

        if tipo_constraint == "maior_que_zero":
            # Obrigatorio: valor > 0
            if valor is None:
                return False, None, f"{col_name} e obrigatorio (> 0)", False
            try:
                num = float(valor) if isinstance(valor, str) else valor
                if num <= 0:
                    return False, None, f"{col_name}={valor} deve ser > 0", False
            except (ValueError, TypeError):
                return False, None, f"{col_name}='{valor}' nao e numero valido", False
            return True, valor, None, False

        elif tipo_constraint == "maior_que_zero_ou_null":
            # Corrigivel: valor > 0 OU NULL
            if valor is None:
                return True, None, None, False
            try:
                num = float(valor) if isinstance(valor, str) else valor
                if num <= 0:
                    return False, None, f"{col_name}={valor} deve ser > 0 ou vazio (sera NULL)", True
            except (ValueError, TypeError):
                # Tentar extrair apenas digitos (ex: "(55)" -> 55)
                if isinstance(valor, str):
                    digitos = ''.join(c for c in valor if c.isdigit())
                    if digitos:
                        try:
                            num = int(digitos)
                            if num > 0:
                                return True, num, None, False  # Retorna valor corrigido
                        except ValueError:
                            pass
                return False, None, f"{col_name}='{valor}' nao e numero valido", True
            return True, valor, None, False

        elif tipo_constraint == "maior_igual_zero":
            # Obrigatorio: valor >= 0
            if valor is None:
                return False, None, f"{col_name} e obrigatorio (>= 0)", False
            try:
                num = float(valor) if isinstance(valor, str) else valor
                if num < 0:
                    return False, None, f"{col_name}={valor} deve ser >= 0", False
            except (ValueError, TypeError):
                return False, None, f"{col_name}='{valor}' nao e numero valido", False
            return True, valor, None, False

        elif tipo_constraint == "maior_igual_zero_ou_null":
            # Corrigivel: valor >= 0 OU NULL
            if valor is None:
                return True, None, None, False
            try:
                num = float(valor) if isinstance(valor, str) else valor
                if num < 0:
                    return False, None, f"{col_name}={valor} deve ser >= 0 ou vazio", True
            except (ValueError, TypeError):
                return False, None, f"{col_name}='{valor}' nao e numero valido", True
            return True, valor, None, False

        elif tipo_constraint == "nao_vazio":
            # Obrigatorio: string nao vazia
            if valor is None or (isinstance(valor, str) and valor.strip() == ""):
                return False, None, f"{col_name} e obrigatorio (nao pode ser vazio)", False
            return True, valor, None, False

        elif tipo_constraint == "nao_vazio_ou_null":
            # Corrigivel: string nao vazia OU NULL
            if valor is None:
                return True, None, None, False
            if isinstance(valor, str) and valor.strip() == "":
                return True, None, None, False  # Vazio -> NULL, OK
            return True, valor, None, False

        elif tipo_constraint == "formato_cnpjcpf":
            valido, corrigido, msg = self._validar_cnpjcpf_formato(valor)
            return valido, corrigido if valido else None, msg, True  # Corrigivel -> NULL

        elif tipo_constraint == "formato_cep":
            valido, corrigido, msg = self._validar_cep_formato(valor)
            return valido, corrigido if valido else None, msg, True  # Corrigivel -> NULL

        elif tipo_constraint == "uf_valida":
            if valor is None:
                return True, None, None, False
            val_upper = str(valor).strip().upper()
            if val_upper == "":
                return True, None, None, False
            if val_upper not in self._UFS_VALIDAS:
                return False, None, f"UF '{valor}' invalida (validas: AC,AL,AM,...,TO)", True
            return True, val_upper, None, False

        elif tipo_constraint == "sim_nao":
            if valor is None:
                return True, None, None, False
            val_upper = str(valor).strip().upper()
            if val_upper == "":
                return True, None, None, False
            if val_upper not in ('S', 'N'):
                return False, None, f"{col_name}='{valor}' deve ser 'S' ou 'N'", True
            return True, val_upper, None, False

        # Constraint desconhecida - passar
        return True, valor, None, False

    def _validar_batch_pre_importacao(self, nome_aba, linhas, header_map, constraints, pk_col):
        """
        Valida todas as linhas antes do batch insert.
        Retorna (erros_bloqueantes, avisos, correcoes).
        - erros_bloqueantes: lista de (linha_excel, pk, col, msg) - impedem importacao
        - avisos: lista de (linha_excel, pk, col, msg, valor_original) - serao corrigidos para NULL
        - correcoes: dict {(linha_idx, col_name): None} - valores a serem corrigidos
        """
        erros_bloqueantes = []
        avisos = []
        correcoes = {}

        pk_idx = header_map.get(pk_col)

        for linha_idx, row in enumerate(linhas):
            # Linha na planilha (considerando header na linha 1)
            linha_excel = linha_idx + 2

            # Obter PK para identificacao
            pk_val = row[pk_idx] if pk_idx is not None and pk_idx < len(row) else "?"

            for col_name, tipo_constraint in constraints.items():
                col_idx = header_map.get(col_name)
                if col_idx is None:
                    # Coluna nao existe na planilha
                    # Se for obrigatoria (nao_vazio ou maior_que_zero), reportar
                    if tipo_constraint in ("nao_vazio", "maior_que_zero", "maior_igual_zero"):
                        # Verificar se e a PK (que ja foi validada antes)
                        if col_name != pk_col:
                            erros_bloqueantes.append((linha_excel, pk_val, col_name, f"Coluna '{col_name}' nao encontrada na planilha"))
                    continue

                valor = row[col_idx] if col_idx < len(row) else None

                valido, valor_corrigido, msg, corrigivel = self._validar_constraint(valor, tipo_constraint, col_name)

                if not valido:
                    if corrigivel:
                        # Valor invalido mas corrigivel -> sera NULL
                        avisos.append((linha_excel, pk_val, col_name, msg, valor))
                        correcoes[(linha_idx, col_name)] = None
                    else:
                        # Erro bloqueante
                        erros_bloqueantes.append((linha_excel, pk_val, col_name, msg))

        return erros_bloqueantes, avisos, correcoes

    def _truncar_valor(self, valor, col_name):
        """Trunca string se exceder o tamanho maximo definido para a coluna."""
        if valor is None:
            return None
        if not isinstance(valor, str):
            return valor
        max_len = self._MAX_LENGTHS.get(col_name)
        if max_len and len(valor) > max_len:
            return valor[:max_len]
        return valor

    def _formatar_cep(self, valor):
        """
        Formata CEP para formato SRPP (NNNNN-NNN, 9 chars) ou None.
        Aceita CEPs com 5-8 digitos (faz pad com zeros a esquerda).
        Retorna NULL se nao puder produzir formato valido.
        """
        if valor is None:
            return None
        val = str(valor).strip()
        if not val:
            return None
        # Ja formatado 9 chars (NNNNN-NNN) - validar se e apenas digitos e hifen
        if len(val) == 9 and val[5] == '-':
            # Verificar se os outros caracteres sao digitos
            parte1 = val[0:5]
            parte2 = val[6:9]
            if parte1.isdigit() and parte2.isdigit():
                return val
            # Formato correto mas caracteres invalidos -> NULL
            return None
        # Extrair apenas digitos
        digitos = ''.join(c for c in val if c.isdigit())
        # Se nao tem digitos ou so tem "0", retorna NULL
        if not digitos or (len(digitos) == 1 and digitos == "0"):
            return None
        # CEP precisa ter entre 5 e 8 digitos para ser valido
        # Menos de 5: nao e CEP valido -> NULL
        if len(digitos) < 5:
            return None
        # Mais de 8: provavelmente nao e CEP (pode ser telefone, CNPJ, etc) -> NULL
        if len(digitos) > 8:
            return None
        # Pad para 8 digitos e formatar
        digitos = digitos.zfill(8)
        return f"{digitos[0:5]}-{digitos[5:8]}"

    def _bulk_insert_staging(self, sql_insert, batch_rows, col_names, pk_col_name="?",
                              staging_table=None, col_types=None):
        """
        Tenta bulk insert (fast_executemany). Se falhar, faz fallback row-by-row.
        Compativel com insercao em chunks (nao trunca staging inteira no fallback).
        """
        import pyodbc as _pyodbc

        # Guardar max identity antes de inserir (para limpar apenas este chunk no fallback)
        identity_before = 0
        if staging_table:
            c = self.conn.cursor()
            try:
                c.execute(f"SELECT ISNULL(MAX(linha), 0) FROM {staging_table}")
                identity_before = c.fetchone()[0]
            finally:
                c.close()

        cursor = self.conn.cursor()
        if self._fast_executemany_ok:
            cursor.fast_executemany = True
            if col_types:
                _type_map = {
                    "datetime": (_pyodbc.SQL_TYPE_TIMESTAMP, 23, 3),
                    "decimal":  (_pyodbc.SQL_DECIMAL, 8, 2),
                    "int":      (_pyodbc.SQL_INTEGER, 0, 0),
                    "smallint": (_pyodbc.SQL_SMALLINT, 0, 0),
                    "tinyint":  (_pyodbc.SQL_TINYINT, 0, 0),
                    "char":     (_pyodbc.SQL_VARCHAR, 10, 0),
                    "varchar":  (_pyodbc.SQL_VARCHAR, 200, 0),
                }
                input_sizes = [_type_map.get(t, (_pyodbc.SQL_VARCHAR, 200, 0)) for t in col_types]
                cursor.setinputsizes(input_sizes)

        try:
            cursor.executemany(sql_insert, batch_rows)
            cursor.close()
            return len(batch_rows), 0
        except Exception as bulk_err:
            cursor.close()
            self._log(f"  Bulk insert falhou: {bulk_err}")

            # Limpar apenas as linhas deste chunk (parciais do executemany que falhou)
            if staging_table:
                c = self.conn.cursor()
                try:
                    c.execute(f"DELETE FROM {staging_table} WHERE linha > {identity_before}")
                    while c.nextset():
                        pass
                finally:
                    c.close()

            self._log(f"  Fazendo fallback row-by-row...")
            inseridos = 0
            erros = 0
            pk_idx = col_names.index(pk_col_name) if pk_col_name in col_names else 0

            for i, row_vals in enumerate(batch_rows):
                pk_val = row_vals[pk_idx] if pk_idx < len(row_vals) else "?"
                cursor = self.conn.cursor()
                try:
                    cursor.execute(sql_insert, row_vals)
                    inseridos += 1
                except Exception as row_err:
                    erros += 1
                    cols_problema = []
                    for j, (col, val) in enumerate(zip(col_names, row_vals)):
                        if val is not None and not isinstance(val, (int, float)):
                            cols_problema.append(f"{col}={repr(val)}")
                    detalhes = ", ".join(cols_problema[:5]) if cols_problema else "valores desconhecidos"
                    self._log(f"  ERRO STAGING | PK=[{pk_val}]: {row_err} | {detalhes}")
                finally:
                    cursor.close()

            self._log(f"  Fallback: {inseridos} inseridos, {erros} com erro de {len(batch_rows)} linhas")
            return inseridos, erros

    def _exec_sem_params(self, nome_proc):
        """Executa procedure sem parametros."""
        cursor = self.conn.cursor()
        try:
            cursor.execute(f"SET NOCOUNT ON; EXEC {nome_proc}")
            # Consumir todos os result sets para capturar erros deferidos
            while cursor.nextset():
                pass
            self._log(f"  OK: {nome_proc}")
            return True, ""
        except pyodbc.Error as e:
            msg = self._extrair_msg_erro(e)
            self._log(f"  ERRO: {nome_proc} - {msg}")
            return False, msg
        finally:
            cursor.close()

    def _exec_com_output(self, nome_proc, params):
        """
        Executa procedure com @msgretorno OUTPUT.
        params: lista de (nome_param, valor)
        Retorna (sucesso, msgretorno).
        """
        cursor = self.conn.cursor()
        try:
            placeholders = ", ".join(f"{nome}=?" for nome, _ in params)
            sql = (
                f"SET NOCOUNT ON; "
                f"DECLARE @msgretorno varchar(250); "
                f"EXEC {nome_proc} @msgretorno=@msgretorno OUTPUT, {placeholders}; "
                f"SELECT @msgretorno;"
            )
            valores = [v for _, v in params]
            cursor.execute(sql, valores)
            row = cursor.fetchone()
            return True, (row[0] if row else "")
        except pyodbc.Error as e:
            return False, self._extrair_msg_erro(e)
        finally:
            cursor.close()

    @staticmethod
    def _extrair_msg_erro(exc):
        """Extrai mensagem legivel de um pyodbc.Error."""
        msg = str(exc)
        # pyodbc retorna formato: "[codigo][driver] mensagem real"
        if "]" in msg:
            msg = msg.split("]")[-1].strip()
        return msg

    # ----------------------------------------------------------
    # Pre-validacao: Tipo de Codigo (EMPRESA vs PRODUTOS)
    # ----------------------------------------------------------

    def _ler_config_tipo_codigo(self, wb):
        """
        Le a configuracao de tipo/tamanho de codigo da aba EMPRESA.
        Celula C7: TipoCodProduto (N=Numerico, A=Alfanumerico)
        Celula C8: TamanhoCodProduto
        Celula C10: TipoCodAuxiliarProduto (N=Numerico, A=Alfanumerico, X=Nao Usado)
        Celula C11: TamanhoCodAuxiliarProduto
        Retorna (tipo_cod, tamanho_cod, tipo_aux, tamanho_aux), com None onde nao encontrar.
        """
        if "EMPRESA" not in wb.sheetnames:
            return None, None, None, None

        sheet = wb["EMPRESA"]
        tipo_cod = None
        tamanho_cod = None
        tipo_aux = None
        tamanho_aux = None

        # Ler por iter_rows (compativel com read_only=True)
        # C7/C8 = row 7/8 col 2 (0-indexed), C10/C11 = row 10/11 col 2
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=11, values_only=True), 1):
            if i == 7 and len(row) > 2 and row[2]:
                val = str(row[2]).strip().upper()
                if val.startswith("N"):
                    tipo_cod = "N"
                elif val.startswith("A"):
                    tipo_cod = "A"
            elif i == 8 and len(row) > 2 and row[2] is not None:
                try:
                    tamanho_cod = int(row[2])
                except (ValueError, TypeError):
                    pass
            elif i == 10 and len(row) > 2 and row[2]:
                val = str(row[2]).strip().upper()
                if val.startswith("N"):
                    tipo_aux = "N"
                elif val.startswith("A"):
                    tipo_aux = "A"
                elif val.startswith("X"):
                    tipo_aux = "X"
            elif i == 11 and len(row) > 2 and row[2] is not None:
                try:
                    tamanho_aux = int(row[2])
                except (ValueError, TypeError):
                    pass

        return tipo_cod, tamanho_cod, tipo_aux, tamanho_aux

    def _valor_e_numerico(self, valor):
        """Verifica se um valor e puramente numerico (apenas digitos)."""
        if valor is None:
            return True  # NULL e valido para ambos os tipos
        val = str(valor).strip()
        if not val:
            return True  # Vazio e valido
        # Remover espacos e verificar se so tem digitos
        return val.isdigit()

    def _validar_tipos_codigo_produtos(self, wb, tipo_cod, tamanho_cod, tipo_aux, tamanho_aux):
        """
        Valida se os codigos na aba PRODUTOS sao compativeis com a configuracao
        da EMPRESA: tipo (numerico/alfanumerico) e tamanho maximo.
        Usa streaming para nao carregar tudo na memoria.
        Retorna (ok, lista_erros).
        """
        header_map = self._ler_header_aba(wb, "PRODUTOS")
        if header_map is None:
            return True, []

        erros = []
        idx_cod = header_map.get("CodProduto")
        idx_aux = header_map.get("CodAuxiliarProduto")

        sheet = wb["PRODUTOS"]
        for linha_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            linha_excel = linha_idx + 2

            if idx_cod is not None:
                cod_val = row[idx_cod] if idx_cod < len(row) else None
                if cod_val is not None and str(cod_val).strip():
                    cod_str = str(cod_val).strip()
                    if tipo_cod == "N" and not self._valor_e_numerico(cod_val):
                        erros.append(
                            f"Linha {linha_excel}: CodProduto '{cod_val}' e alfanumerico, "
                            f"mas configuracao EMPRESA (C7) diz N=Numerico"
                        )
                    elif tamanho_cod and len(cod_str) > tamanho_cod:
                        erros.append(
                            f"Linha {linha_excel}: CodProduto '{cod_val}' tem {len(cod_str)} caracteres, "
                            f"maior que o TamanhoCodProduto configurado na EMPRESA (C8={tamanho_cod})"
                        )
                    if len(erros) >= 10:
                        erros.append("... (mais erros omitidos)")
                        break

            if idx_aux is not None:
                aux_val = row[idx_aux] if idx_aux < len(row) else None
                if aux_val is not None and str(aux_val).strip():
                    aux_str = str(aux_val).strip()
                    if tipo_aux == "N" and not self._valor_e_numerico(aux_val):
                        erros.append(
                            f"Linha {linha_excel}: CodAuxiliarProduto '{aux_val}' e alfanumerico, "
                            f"mas configuracao EMPRESA (C10) diz N=Numerico"
                        )
                    elif tamanho_aux and len(aux_str) > tamanho_aux:
                        erros.append(
                            f"Linha {linha_excel}: CodAuxiliarProduto '{aux_val}' tem {len(aux_str)} caracteres, "
                            f"maior que o TamanhoCodAuxiliarProduto configurado na EMPRESA (C11={tamanho_aux})"
                        )
                    if len(erros) >= 10:
                        erros.append("... (mais erros omitidos)")
                        break

        return len(erros) == 0, erros

    # ----------------------------------------------------------
    # Leitura da planilha
    # ----------------------------------------------------------

    def _ler_header_aba(self, wb, nome_aba):
        """Le apenas o header de uma aba. Compativel com read_only=True."""
        if nome_aba not in wb.sheetnames:
            return None
        sheet = wb[nome_aba]
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            return {name: idx for idx, name in enumerate(row) if name is not None}
        return None

    def _ler_linhas_aba(self, wb, nome_aba):
        """Retorna (header_map, rows) de uma aba. Usado apenas para abas pequenas."""
        header_map = self._ler_header_aba(wb, nome_aba)
        if header_map is None:
            return None, []
        sheet = wb[nome_aba]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rows.append(row)
        return header_map, rows

    # ----------------------------------------------------------
    # Importacao: EMPRESA (formato especial chave-valor)
    # ----------------------------------------------------------

    def _importar_empresa(self, wb, sobreescreve):
        t0 = time.time()
        self._log("--- Importando EMPRESA ---")
        config = MAPA_ABAS["EMPRESA"]

        if "EMPRESA" not in wb.sheetnames:
            self._log("  Aba EMPRESA nao encontrada!")
            return 0, 1

        sheet = wb["EMPRESA"]
        linhas = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                codigo = int(row[0])
            except (ValueError, TypeError):
                continue
            descricao = str(row[1]).strip() if row[1] is not None else ""
            configuracao = str(row[2]).strip() if row[2] is not None else ""
            linhas.append((codigo, descricao, configuracao))

        if not linhas:
            self._log("  Nenhuma linha de configuracao encontrada")
            return 0, 1

        total = len(linhas)
        erros = 0
        sucesso = 0

        # Passo 1: limpar
        self._exec_sem_params(config["pkunica1"])

        # Passo 2: PKUnica2
        self._log(f"  Registrando {total} codigos...")
        for codigo, _, _ in linhas:
            if self.cancelar:
                return sucesso, erros
            params = [("@sobreescreve", sobreescreve), ("@codigo", codigo)]
            ok, msg = self._exec_com_output(config["pkunica2"], params)
            if not ok:
                self._log(f"  ERRO PKUnica2 cod={codigo}: {msg}")
                erros += 1

        # Passo 3: importar cada parametro
        self._log(f"  Importando {total} parametros...")
        for i, (codigo, descricao, configuracao) in enumerate(linhas):
            if self.cancelar:
                return sucesso, erros
            params = [
                ("@sobreescreve", sobreescreve),
                ("@Codigo", codigo), ("@Codigo_rec", 1),
                ("@Descricao", descricao), ("@Descricao_rec", 1),
                ("@Configuracao", configuracao), ("@Configuracao_rec", 1),
            ]
            ok, msg = self._exec_com_output(config["procedure"], params)
            if ok:
                sucesso += 1
            else:
                erros += 1
                self._log(f"  ERRO cod={codigo}: {msg}")

        # Passo 4: ConfigurarEmpresa
        self._log("  Executando ConfigurarEmpresa...")
        cursor = self.conn.cursor()
        try:
            sql = "DECLARE @msgretorno varchar(250); EXEC ConfigurarEmpresa @msgretorno=@msgretorno OUTPUT; SELECT @msgretorno;"
            cursor.execute(sql)
            row = cursor.fetchone()
            self._log(f"  ConfigurarEmpresa: {row[0] if row else 'OK'}")
        except pyodbc.Error as e:
            erros += 1
            self._log(f"  ERRO ConfigurarEmpresa: {self._extrair_msg_erro(e)}")
        finally:
            cursor.close()

        elapsed = time.time() - t0
        self._log(f"  EMPRESA: {sucesso} ok, {erros} erros | {total} linhas em {elapsed:.1f}s")
        return sucesso, erros

    # ----------------------------------------------------------
    # Importacao: aba padrao (fluxo PKUnica1 -> PKUnica2 -> Proc)
    # ----------------------------------------------------------

    def _importar_aba_padrao(self, wb, nome_aba, sobreescreve, prog_base, prog_range):
        t0 = time.time()
        self._log(f"--- Importando {nome_aba} ---")
        config = MAPA_ABAS[nome_aba]
        colunas = config["colunas"]
        pk_cols = config["pk_columns"]

        header_map, rows = self._ler_linhas_aba(wb, nome_aba)
        if header_map is None:
            self._log(f"  Aba {nome_aba} nao encontrada!")
            return 0, 1

        # Filtrar linhas com PK preenchida
        pk0_name = pk_cols[0][0]
        pk0_idx = header_map.get(pk0_name)
        if pk0_idx is None:
            self._log(f"  Coluna PK '{pk0_name}' nao encontrada em {nome_aba}")
            return 0, 1

        linhas = [r for r in rows if pk0_idx < len(r) and r[pk0_idx] is not None and str(r[pk0_idx]).strip()]
        total = len(linhas)
        if total == 0:
            self._log(f"  Nenhuma linha valida em {nome_aba}")
            return 0, 0

        erros = 0
        sucesso = 0

        # Passo 1: limpar
        self._exec_sem_params(config["pkunica1"])

        # Passo 2: PKUnica2
        self._progresso(prog_base, f"Registrando PKs de {nome_aba}...")
        for row in linhas:
            if self.cancelar:
                return sucesso, erros
            params = [("@sobreescreve", sobreescreve)]
            for col_excel, param_sql, tipo_sql in pk_cols:
                idx = header_map.get(col_excel)
                val = row[idx] if idx is not None and idx < len(row) else None
                params.append((param_sql, self._converter_valor(val, tipo_sql)))
            ok, msg = self._exec_com_output(config["pkunica2"], params)
            if not ok:
                pk_str = self._pk_str(row, pk_cols, header_map)
                self._log(f"  ERRO PKUnica2 {nome_aba} PK=[{pk_str}]: {msg}")
                erros += 1

        # Passo 3: importacao principal
        for i, row in enumerate(linhas):
            if self.cancelar:
                return sucesso, erros

            if total > 0:
                pct = prog_base + int((i / total) * prog_range)
                self._progresso(pct, f"Importando {nome_aba}... {i+1}/{total}")

            params = [("@sobreescreve", sobreescreve)]
            for col_excel, param_sql, tipo_sql in colunas:
                idx = header_map.get(col_excel)
                if idx is not None and idx < len(row):
                    valor = self._converter_valor(row[idx], tipo_sql)
                    if param_sql.lower() == "@cnpjcpf":
                        valor = self._formatar_cnpjcpf(valor)
                    elif param_sql.lower() == "@cep":
                        valor = self._formatar_cep(valor)
                    # Campos com constraint > 0 OR NULL: converter 0 para NULL
                    if param_sql.lower() in self._ZERO_TO_NULL and valor == 0:
                        valor = None
                    # Strings vazias -> NULL (constraints NaoVazioNull)
                    if isinstance(valor, str) and valor.strip() == "":
                        valor = None
                    # Truncar se exceder tamanho maximo
                    valor = self._truncar_valor(valor, col_excel)
                    rec = 1  # coluna mapeada
                else:
                    valor = None
                    rec = None  # coluna nao mapeada
                params.append((param_sql, valor))
                params.append((f"{param_sql}_rec", rec))

            ok, msg = self._exec_com_output(config["procedure"], params)
            if ok:
                sucesso += 1
            else:
                erros += 1
                pk_str = self._pk_str(row, pk_cols, header_map)
                self._log(f"  ERRO {nome_aba} linha {i+1} PK=[{pk_str}]: {msg}")

        elapsed = time.time() - t0
        self._log(f"  {nome_aba}: {sucesso} ok, {erros} erros | {total} linhas em {elapsed:.1f}s")
        return sucesso, erros

    # ----------------------------------------------------------
    # Importacao em lote: PRODUTOS (staging + procedure batch)
    # ----------------------------------------------------------

    def _setup_batch_objects(self):
        """Cria staging table e procedure de lote se nao existirem."""
        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_STAGING)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_PROCEDURE)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_PROCEDURE_BODY)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

    _CHUNK_SIZE = 10000

    def _importar_produtos_lote(self, wb, sobreescreve, prog_base, prog_range):
        """Importa PRODUTOS usando staging table + procedure batch. Streaming em chunks."""
        t0 = time.time()
        self._log("--- Importando PRODUTOS (modo lote) ---")

        config = MAPA_ABAS["PRODUTOS"]
        colunas = config["colunas"]

        header_map = self._ler_header_aba(wb, "PRODUTOS")
        if header_map is None:
            self._log("  Aba PRODUTOS nao encontrada!")
            return 0, 1

        pk0_idx = header_map.get("CodProduto")
        if pk0_idx is None:
            self._log("  Coluna PK 'CodProduto' nao encontrada em PRODUTOS")
            return 0, 1

        # 1. Criar objetos SQL
        self._progresso(prog_base, "Preparando importacao em lote...")
        self._setup_batch_objects()

        # 2. Limpar staging
        cursor = self.conn.cursor()
        try:
            cursor.execute("TRUNCATE TABLE ImportaProdutoAmbos_Staging")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute("DBCC CHECKIDENT ('ImportaProdutoAmbos_Staging', RESEED, 0)")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        # 3. Ler configuracao de tamanho de codigo do banco para zero-padding
        tamanho_cod = None
        tamanho_cod_aux = None
        try:
            c = self.conn.cursor()
            c.execute(
                "SELECT TipoCodProduto, TamanhoCodProduto, TipoCodAuxProduto, TamanhoCodAuxProduto "
                "FROM Empresa WHERE CodEmpresa = 1"
            )
            row_emp = c.fetchone()
            c.close()
            if row_emp and row_emp[0] == "N" and row_emp[1]:
                tamanho_cod = int(row_emp[1])
                self._log(f"  TipoCodProduto=N, TamanhoCodProduto={tamanho_cod} (zero-padding ativo)")
            if row_emp and row_emp[2] == "N" and row_emp[3]:
                tamanho_cod_aux = int(row_emp[3])
                self._log(f"  TipoCodAuxProduto=N, TamanhoCodAuxProduto={tamanho_cod_aux} (zero-padding ativo)")
        except Exception as e:
            self._log(f"  AVISO: nao foi possivel ler TamanhoCodProduto/TamanhoCodAuxProduto: {e}")

        # 4. Streaming em chunks: ler da planilha e inserir no staging em lotes
        col_order = [
            "CodProduto", "CodAuxiliarProduto", "Produto",
            "PrecoTabela1", "PrecoTabela2", "PrecoTabela3",
            "QtdeEstoqueAtual", "QtdeEstoqueFuturo", "DtEstoqueFuturo",
            "LimiteDescIndividual", "AliquotaIPI", "MultiploGrade", "DescontoGrade",
            "PathFotografia", "CodFilial", "PrecoPromocional", "TipoVendaSemEstoque",
            "CodFamilia", "CodEstilo", "QtdeMultipla", "QtdeMinima",
            "QtdeTabela1", "QtdeTabela2", "QtdeTabela3", "QtdeEtiquetas",
        ]
        col_type = {c[0]: c[2] for c in colunas}
        col_sql_param = {c[0]: c[1].lower() for c in colunas}
        col_types_list = [col_type.get(c, "varchar") for c in col_order]

        total = 0
        total_inseridos = 0
        total_erros_staging = 0
        chunk = []

        sheet = wb["PRODUTOS"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            if pk0_idx >= len(row) or row[pk0_idx] is None or not str(row[pk0_idx]).strip():
                continue

            total += 1
            vals = []
            for col_name in col_order:
                idx = header_map.get(col_name)
                if idx is not None and idx < len(row):
                    valor = self._converter_valor(row[idx], col_type.get(col_name, "varchar"))
                    if isinstance(valor, str) and valor.strip() == "":
                        valor = None
                    # Zero-padding: CodProduto/CodAuxiliarProduto numerico pode perder zeros a esquerda no openpyxl
                    if col_name == "CodProduto" and tamanho_cod and isinstance(valor, str):
                        digitos = ''.join(c for c in valor if c.isdigit())
                        if digitos == valor and len(valor) < tamanho_cod:
                            valor = valor.zfill(tamanho_cod)
                    elif col_name == "CodAuxiliarProduto" and tamanho_cod_aux and isinstance(valor, str):
                        digitos = ''.join(c for c in valor if c.isdigit())
                        if digitos == valor and len(valor) < tamanho_cod_aux:
                            valor = valor.zfill(tamanho_cod_aux)
                    param_name = col_sql_param.get(col_name, "")
                    if param_name in self._ZERO_TO_NULL and valor == 0:
                        valor = None
                    valor = self._truncar_valor(valor, col_name)
                else:
                    valor = None
                vals.append(valor)
            chunk.append(vals)

            if len(chunk) >= self._CHUNK_SIZE:
                self._progresso(prog_base + 2, f"Inserindo no staging... ({total} linhas)")
                ins, err = self._bulk_insert_staging(
                    SQL_INSERT_STAGING, chunk, col_order, pk_col_name="CodProduto",
                    staging_table="ImportaProdutoAmbos_Staging", col_types=col_types_list
                )
                total_inseridos += ins
                total_erros_staging += err
                chunk = []

        # Ultimo chunk
        if chunk:
            self._progresso(prog_base + 2, f"Inserindo no staging... ({total} linhas)")
            ins, err = self._bulk_insert_staging(
                SQL_INSERT_STAGING, chunk, col_order, pk_col_name="CodProduto",
                staging_table="ImportaProdutoAmbos_Staging", col_types=col_types_list
            )
            total_inseridos += ins
            total_erros_staging += err

        if total == 0:
            self._log("  Nenhuma linha valida em PRODUTOS")
            return 0, 0

        if total_erros_staging > 0:
            self._log(f"  {total_inseridos} linhas inseridas no staging ({total_erros_staging} com erro).")
        else:
            self._log(f"  {total} linhas inseridas no staging.")

        self._progresso(prog_base + int(prog_range * 0.3), "Executando validacao e importacao em lote...")

        # 5. Chamar a procedure batch
        cursor = self.conn.cursor()
        try:
            cursor.execute(f"SET NOCOUNT ON; EXEC ImportaProdutoAmbos_Lote @sobreescreve={sobreescreve}")
            results = cursor.fetchall()
        finally:
            cursor.close()

        # 6. Processar resultados
        sucesso = 0
        erros = 0
        for row in results:
            linha_num = row[0]
            codprod = row[1]
            codaux = row[2]
            status = row[3]
            mensagem = row[4]

            if status == "OK":
                sucesso += 1
            else:
                erros += 1
                self._log(
    f"  ERRO PRODUTOS | DB={self.database_name} | PROC=ImportaProdutoAmbos_Lote "
    f"| linha {linha_num} PK=[{codprod}, {codaux}]: {mensagem}"
)


        # Limpar staging table após processamento
        cursor = self.conn.cursor()
        try:
            cursor.execute("TRUNCATE TABLE ImportaProdutoAmbos_Staging")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        elapsed = time.time() - t0
        self._log(f"  PRODUTOS: {sucesso} ok, {erros} erros | {total} linhas em {elapsed:.1f}s")
        self._progresso(prog_base + prog_range, f"PRODUTOS concluido: {sucesso} ok, {erros} erros")
        return sucesso, erros

    # ----------------------------------------------------------
    # Importacao em lote: TRANSPORTADORA
    # ----------------------------------------------------------

    def _setup_batch_objects_transportadora(self):
        """Cria staging table e procedure de lote para transportadora."""
        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_STAGING_TRANSPORTADORA)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_PROCEDURE_TRANSPORTADORA)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_PROCEDURE_BODY_TRANSPORTADORA)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

    def _importar_transportadoras_lote(self, wb, sobreescreve, prog_base, prog_range):
        """Importa TRANSP usando staging table + procedure batch. Streaming em chunks."""
        t0 = time.time()
        self._log("--- Importando TRANSP (modo lote) ---")

        config = MAPA_ABAS["TRANSP"]
        header_map = self._ler_header_aba(wb, "TRANSP")
        if header_map is None:
            self._log("  Aba TRANSP nao encontrada!")
            return 0, 1

        pk0_idx = header_map.get("CodTransportadora")
        if pk0_idx is None:
            self._log("  Coluna PK 'CodTransportadora' nao encontrada em TRANSP")
            return 0, 1

        # 1. Setup
        self._progresso(prog_base, "Preparando importacao em lote (TRANSP)...")
        self._setup_batch_objects_transportadora()

        # 2. Limpar staging
        cursor = self.conn.cursor()
        try:
            cursor.execute("TRUNCATE TABLE ImportaTransportadora_Staging")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute("DBCC CHECKIDENT ('ImportaTransportadora_Staging', RESEED, 0)")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        # 3. Streaming em chunks
        colunas = config["colunas"]
        col_order = ["CodTransportadora", "Transportadora", "TransportadoraPadrao"]
        col_type = {c[0]: c[2] for c in colunas}
        col_sql_param = {c[0]: c[1].lower() for c in colunas}
        col_types_list = [col_type.get(c, "varchar") for c in col_order]

        total = 0
        total_inseridos = 0
        total_erros_staging = 0
        chunk = []

        sheet = wb["TRANSP"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            if pk0_idx >= len(row) or row[pk0_idx] is None or not str(row[pk0_idx]).strip():
                continue

            total += 1
            vals = []
            for col_name in col_order:
                idx = header_map.get(col_name)
                if idx is not None and idx < len(row):
                    valor = self._converter_valor(row[idx], col_type.get(col_name, "varchar"))
                    if isinstance(valor, str) and valor.strip() == "":
                        valor = None
                    param_name = col_sql_param.get(col_name, "")
                    if param_name in self._ZERO_TO_NULL and valor == 0:
                        valor = None
                    valor = self._truncar_valor(valor, col_name)
                else:
                    valor = None
                vals.append(valor)
            chunk.append(vals)

            if len(chunk) >= self._CHUNK_SIZE:
                self._progresso(prog_base + 2, f"Inserindo transp no staging... ({total} linhas)")
                ins, err = self._bulk_insert_staging(
                    SQL_INSERT_STAGING_TRANSPORTADORA, chunk, col_order, pk_col_name="CodTransportadora",
                    staging_table="ImportaTransportadora_Staging", col_types=col_types_list
                )
                total_inseridos += ins
                total_erros_staging += err
                chunk = []

        if chunk:
            ins, err = self._bulk_insert_staging(
                SQL_INSERT_STAGING_TRANSPORTADORA, chunk, col_order, pk_col_name="CodTransportadora",
                staging_table="ImportaTransportadora_Staging", col_types=col_types_list
            )
            total_inseridos += ins
            total_erros_staging += err

        if total == 0:
            self._log("  Nenhuma linha valida em TRANSP")
            return 0, 0

        if total_erros_staging > 0:
            self._log(f"  {total_inseridos} linhas inseridas no staging ({total_erros_staging} com erro).")
        else:
            self._log(f"  {total} linhas inseridas no staging.")
        self._progresso(prog_base + int(prog_range * 0.3), "Executando importacao em lote (TRANSP)...")

        # 5. Executar procedure
        cursor = self.conn.cursor()
        try:
            cursor.execute(f"SET NOCOUNT ON; EXEC ImportaTransportadora_Lote @sobreescreve={sobreescreve}")
            results = cursor.fetchall()
        finally:
            cursor.close()

        # 6. Processar resultados
        sucesso = 0
        erros = 0
        for row in results:
            linha_num = row[0]
            codtransp = row[1]
            status = row[2]
            mensagem = row[3]

            if status == "OK":
                sucesso += 1
            else:
                erros += 1
                self._log(f"  ERRO TRANSP | DB={self.database_name} | linha {linha_num} PK=[{codtransp}]: {mensagem}")

        # 7. Limpar staging
        cursor = self.conn.cursor()
        try:
            cursor.execute("TRUNCATE TABLE ImportaTransportadora_Staging")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        elapsed = time.time() - t0
        self._log(f"  TRANSP: {sucesso} ok, {erros} erros | {total} linhas em {elapsed:.1f}s")
        self._progresso(prog_base + prog_range, f"TRANSP concluido: {sucesso} ok, {erros} erros")
        return sucesso, erros

    # ----------------------------------------------------------
    # Importacao em lote: CLIENTE
    # ----------------------------------------------------------

    def _setup_batch_objects_cliente(self):
        """Cria staging table e procedure de lote para cliente."""
        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_STAGING_CLIENTE)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_PROCEDURE_CLIENTE)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute(SQL_CREATE_PROCEDURE_BODY_CLIENTE)
            while cursor.nextset():
                pass
        finally:
            cursor.close()

    def _importar_clientes_lote(self, wb, sobreescreve, prog_base, prog_range):
        """Importa CLIENTES usando staging table + procedure batch."""
        import traceback as tb  # Import local para evitar problemas no PyInstaller
        t0 = time.time()
        self._log("--- Importando CLIENTES (modo lote) ---")

        try:
            return self._importar_clientes_lote_impl(wb, sobreescreve, prog_base, prog_range, t0)
        except Exception as e:
            self._log(f"  ERRO em _importar_clientes_lote: {type(e).__name__}: {e}")
            self._log(f"  Traceback: {tb.format_exc()}")
            raise

    def _importar_clientes_lote_impl(self, wb, sobreescreve, prog_base, prog_range, t0):
        """Implementacao real do _importar_clientes_lote. Streaming em chunks."""

        config = MAPA_ABAS["CLIENTES"]
        header_map = self._ler_header_aba(wb, "CLIENTES")
        if header_map is None:
            self._log("  Aba CLIENTES nao encontrada!")
            return 0, 1

        pk0_idx = header_map.get("CodCliente")
        if pk0_idx is None:
            self._log("  Coluna PK 'CodCliente' nao encontrada em CLIENTES")
            return 0, 1

        # 1. Setup
        self._progresso(prog_base, "Preparando importacao em lote (CLIENTES)...")
        self._setup_batch_objects_cliente()

        # 2. Limpar staging
        cursor = self.conn.cursor()
        try:
            cursor.execute("TRUNCATE TABLE ImportaCliente_Staging")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute("DBCC CHECKIDENT ('ImportaCliente_Staging', RESEED, 0)")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        # 3. Streaming em chunks
        col_order = [
            "CodCliente", "CodRepresentante", "NomeFantasia", "RazaoSocial",
            "CNPJCPF", "IERG", "Logradouro", "Bairro", "Cidade", "UF", "CEP",
            "DDD", "Telefone1", "Telefone2", "FAX", "NomeContato",
            "NomeTransportadora", "Observacao", "EMail", "PrecoTabela", "CodTransportadora"
        ]
        col_type_cli = {c[0]: c[2] for c in config["colunas"]}
        col_types_list = [col_type_cli.get(c, "varchar") for c in col_order]

        total = 0
        total_inseridos = 0
        total_erros_staging = 0
        chunk = []

        sheet = wb["CLIENTES"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            if pk0_idx >= len(row) or row[pk0_idx] is None or not str(row[pk0_idx]).strip():
                continue

            total += 1
            vals = []
            for col_name in col_order:
                idx = header_map.get(col_name)
                if idx is not None and idx < len(row):
                    tipo = "int" if col_name in ("CodCliente", "CodRepresentante", "DDD", "PrecoTabela", "CodTransportadora") else "varchar"
                    valor = self._converter_valor(row[idx], tipo)
                    if isinstance(valor, str) and valor.strip() == "":
                        valor = None
                    if isinstance(valor, str):
                        valor = valor.replace("'", "")
                        if not valor:
                            valor = None
                    if col_name == "CNPJCPF" and valor:
                        valor = self._formatar_cnpjcpf(valor)
                    elif col_name == "CEP" and valor:
                        valor = self._formatar_cep(valor)
                    elif col_name == "DDD" and valor is not None:
                        if isinstance(valor, str):
                            digitos = ''.join(c for c in valor if c.isdigit())
                            valor = int(digitos) if digitos else None
                        if valor == 0:
                            valor = None
                    elif col_name == "UF" and valor is not None:
                        uf_upper = str(valor).strip().upper()
                        if uf_upper not in self._UFS_VALIDAS:
                            valor = None
                        else:
                            valor = uf_upper
                    elif col_name in ("Telefone1", "Telefone2", "FAX") and valor is not None:
                        primeiro = str(valor).split("/")[0].strip()
                        digitos = ''.join(c for c in primeiro if c.isdigit())
                        valor = digitos if digitos else None
                    if col_name in ("CodRepresentante", "CodTransportadora", "PrecoTabela") and valor == 0:
                        valor = None
                    valor = self._truncar_valor(valor, col_name)
                else:
                    valor = None
                vals.append(valor)
            chunk.append(vals)

            if len(chunk) >= self._CHUNK_SIZE:
                self._progresso(prog_base + 2, f"Inserindo clientes no staging... ({total} linhas)")
                ins, err = self._bulk_insert_staging(
                    SQL_INSERT_STAGING_CLIENTE, chunk, col_order, pk_col_name="CodCliente",
                    staging_table="ImportaCliente_Staging", col_types=col_types_list
                )
                total_inseridos += ins
                total_erros_staging += err
                chunk = []

        if chunk:
            self._progresso(prog_base + 2, f"Inserindo clientes no staging... ({total} linhas)")
            ins, err = self._bulk_insert_staging(
                SQL_INSERT_STAGING_CLIENTE, chunk, col_order, pk_col_name="CodCliente",
                staging_table="ImportaCliente_Staging", col_types=col_types_list
            )
            total_inseridos += ins
            total_erros_staging += err

        if total == 0:
            self._log("  Nenhuma linha valida em CLIENTES")
            return 0, 0

        if total_erros_staging > 0:
            self._log(f"  {total_inseridos} linhas inseridas no staging ({total_erros_staging} com erro).")
        else:
            self._log(f"  {total} linhas inseridas no staging.")
        self._progresso(prog_base + int(prog_range * 0.3), "Executando importacao em lote (CLIENTES)...")

        # 5. Executar procedure
        cursor = self.conn.cursor()
        try:
            cursor.execute(f"SET NOCOUNT ON; EXEC ImportaCliente_Lote @sobreescreve={sobreescreve}")
            results = cursor.fetchall()
        finally:
            cursor.close()

        # 6. Processar resultados
        sucesso = 0
        erros = 0
        for row in results:
            linha_num = row[0]
            codcliente = row[1]
            status = row[2]
            mensagem = row[3]

            if status == "OK":
                sucesso += 1
            else:
                erros += 1
                self._log(f"  ERRO CLIENTES | DB={self.database_name} | linha {linha_num} PK=[{codcliente}]: {mensagem}")

        # 7. Limpar staging
        cursor = self.conn.cursor()
        try:
            cursor.execute("TRUNCATE TABLE ImportaCliente_Staging")
            while cursor.nextset():
                pass
        finally:
            cursor.close()

        elapsed = time.time() - t0
        self._log(f"  CLIENTES: {sucesso} ok, {erros} erros | {total} linhas em {elapsed:.1f}s")
        self._progresso(prog_base + prog_range, f"CLIENTES concluido: {sucesso} ok, {erros} erros")
        return sucesso, erros

    def _pk_str(self, row, pk_cols, header_map):
        """Monta string com valores das PKs para log."""
        vals = []
        for col_excel, _, _ in pk_cols:
            idx = header_map.get(col_excel)
            val = row[idx] if idx is not None and idx < len(row) else "?"
            vals.append(str(val))
        return ", ".join(vals)

    # ----------------------------------------------------------
    # Backup
    # ----------------------------------------------------------

    def _fazer_backup(self, operacao="Apagar Pedidos e Cadastros e Configuracoes"):
        """
        Faz backup do banco antes de operacoes destrutivas.
        Salva no caminho padrao da instancia SQL Server.
        Formato: {BANCO} - {EMPRESA} - {DATA} {HORA} - {OPERACAO}.bak
        """
        self._log("Fazendo backup do banco de dados...")

        backup_path = None
        cursor = self.conn.cursor()
        try:
            # 1. Tentar obter caminho de backup da instancia (SQL 2012+)
            cursor.execute("SELECT SERVERPROPERTY('InstanceDefaultBackupPath')")
            row = cursor.fetchone()
            backup_path = row[0] if row and row[0] else None
        except Exception:
            pass
        finally:
            cursor.close()

        # Fallback: usar diretorio do arquivo de dados do banco
        if not backup_path:
            cursor = self.conn.cursor()
            try:
                cursor.execute(f"""
                    SELECT LEFT(physical_name, LEN(physical_name) - CHARINDEX('\\', REVERSE(physical_name)))
                    FROM sys.master_files
                    WHERE database_id = DB_ID('{self.database_name}') AND type = 0
                """)
                row = cursor.fetchone()
                backup_path = row[0] if row else None
            except Exception:
                pass
            finally:
                cursor.close()

        if not backup_path:
            self._log("  AVISO: Nao foi possivel obter caminho de backup")
            return False

        cursor = self.conn.cursor()
        try:
            # 2. Buscar nome da empresa
            cursor.execute("SELECT Empresa FROM empresa WHERE codempresa = 1")
            row = cursor.fetchone()
            empresa_nome = row[0] if row else "SEM_EMPRESA"
            # Limpar caracteres invalidos para nome de arquivo
            empresa_nome = empresa_nome.replace("/", "-").replace("\\", "-").replace(":", "-")
            empresa_nome = empresa_nome.replace("*", "-").replace("?", "-").replace("\"", "-")
            empresa_nome = empresa_nome.replace("<", "-").replace(">", "-").replace("|", "-")
        finally:
            cursor.close()

        # 3. Gerar nome do arquivo
        agora = datetime.now()
        data_hora = agora.strftime("%Y-%m-%d %Hh%Mm%S")
        nome_arquivo = f"{self.database_name} - {empresa_nome} - {data_hora} - {operacao}.bak"
        caminho_completo = os.path.join(backup_path, nome_arquivo)

        self._log(f"  Destino: {caminho_completo}")

        cursor = self.conn.cursor()
        try:
            # 4. Executar backup
            sql = f"BACKUP DATABASE [{self.database_name}] TO DISK = N'{caminho_completo}' WITH NOFORMAT, INIT, NAME = N'{self.database_name}-Full', SKIP, NOREWIND, NOUNLOAD, STATS = 10"
            cursor.execute(sql)
            # Consumir todos os result sets (mensagens de progresso)
            while cursor.nextset():
                pass
            self._log(f"  Backup concluido com sucesso!")
            return True
        except pyodbc.Error as e:
            self._log(f"  ERRO no backup: {self._extrair_msg_erro(e)}")
            return False
        finally:
            cursor.close()

    # ----------------------------------------------------------
    # Limpeza
    # ----------------------------------------------------------

    def excluir_tudo(self):
        # Fazer backup antes de excluir
        self._fazer_backup("Apagar Pedidos e Cadastros e Configuracoes")

        # Limpar tabelas que a procedure nao deleta mas que tem FK para as que ela deleta
        self._log("Limpando tabelas com FK dependente...")
        cursor = self.conn.cursor()
        try:
            cursor.execute("SET NOCOUNT ON; DELETE FROM RestricaoCondicaoPagamento")
            while cursor.nextset():
                pass
            self._log("  OK: RestricaoCondicaoPagamento limpa")
        except pyodbc.Error as e:
            self._log(f"  AVISO: RestricaoCondicaoPagamento - {self._extrair_msg_erro(e)}")
        finally:
            cursor.close()

        cursor = self.conn.cursor()
        try:
            cursor.execute("SET NOCOUNT ON; DELETE FROM Representantes_X_AspNetUsers")
            while cursor.nextset():
                pass
            self._log("  OK: Representantes_X_AspNetUsers limpa")
        except pyodbc.Error as e:
            self._log(f"  AVISO: Representantes_X_AspNetUsers - {self._extrair_msg_erro(e)}")
        finally:
            cursor.close()

        self._log("Executando ExcluiPedidosCadastrosConfiguracao...")
        ok, msg = self._exec_sem_params("ExcluiPedidosCadastrosConfiguracao")
        if ok:
            self._log("Todos os dados foram excluidos.")
        return ok

    def limpar_auxiliares(self, abas):
        self._log("Limpando tabelas auxiliares...")
        for aba in abas:
            config = MAPA_ABAS.get(aba)
            if config:
                self._exec_sem_params(config["pkunica1"])

    # ----------------------------------------------------------
    # Fluxo principal
    # ----------------------------------------------------------

    def importar(self, arquivo_excel, abas_selecionadas, sobreescreve,
                 excluir_tudo=False, limpar_auxiliares=False):
        """
        Executa importacao completa.
        Retorna dict {aba: {"sucesso": N, "erros": N}} ou {"ERRO_GERAL": msg}.
        """
        resultados = {}
        self.cancelar = False
        t_total = time.time()

        try:
            self._progresso(0, "Conectando ao SQL Server...")
            self.conectar()
            self._log("Conectado ao SQL Server.")

            # Remove triggers obsoletas de SyncGeral (se existirem no banco)
            cursor = self.conn.cursor()
            try:
                cursor.execute(SQL_REMOVE_SYNC_TRIGGERS)
                while cursor.nextset():
                    pass
            finally:
                cursor.close()

            self._progresso(2, "Carregando planilha...")
            wb = load_workbook(arquivo_excel, data_only=True, read_only=True)
            self._normalizar_nomes_abas(wb)
            self._log(f"Planilha: {os.path.basename(arquivo_excel)}")

            # ============================================================
            # PRE-VALIDACAO CRITICA: Tipo/Tamanho de Codigo (EMPRESA vs PRODUTOS)
            # Se a configuracao diz "Numerico" mas os dados sao alfanumericos,
            # ou se o codigo excede o tamanho configurado, a importacao entra
            # no banco sem erro mas quebra depois (ex: na exportacao). Melhor
            # abortar agora.
            # ============================================================
            if "PRODUTOS" in abas_selecionadas:
                self._progresso(3, "Validando configuracao de tipos de codigo...")
                tipo_cod, tamanho_cod, tipo_aux, tamanho_aux = self._ler_config_tipo_codigo(wb)
                self._log(
                    f"  Config EMPRESA: TipoCodProduto={tipo_cod} (tamanho={tamanho_cod}), "
                    f"TipoCodAuxiliar={tipo_aux} (tamanho={tamanho_aux})"
                )

                if tipo_cod or tipo_aux or tamanho_cod or tamanho_aux:
                    ok, erros_tipo = self._validar_tipos_codigo_produtos(wb, tipo_cod, tamanho_cod, tipo_aux, tamanho_aux)
                    if not ok:
                        self._log("=" * 60)
                        self._log("ERRO CRITICO: Incompatibilidade de tipo/tamanho de codigo!")
                        self._log("=" * 60)
                        for erro in erros_tipo:
                            self._log(f"  {erro}")
                        self._log("")
                        self._log("SOLUCAO: Corrija a configuracao na aba EMPRESA (C7/C8/C10/C11)")
                        self._log("         ou corrija os codigos na aba PRODUTOS.")
                        self._log("=" * 60)
                        wb.close()
                        return {"ERRO_GERAL": "Tipo/tamanho de codigo incompativel entre EMPRESA e PRODUTOS. Veja o log."}

            if excluir_tudo:
                self._progresso(3, "Excluindo todos os dados...")
                if not self.excluir_tudo():
                    return {"ERRO_GERAL": "Falha ao excluir dados"}

            if limpar_auxiliares:
                self._progresso(4, "Limpando tabelas auxiliares...")
                self.limpar_auxiliares(abas_selecionadas)

            abas_ordenadas = [a for a in ORDEM_IMPORTACAO if a in abas_selecionadas]
            total_abas = len(abas_ordenadas)

            for idx, nome_aba in enumerate(abas_ordenadas):
                if self.cancelar:
                    self._log("IMPORTACAO CANCELADA")
                    break

                prog_base = 5 + int((idx / total_abas) * 90)
                prog_range = int(90 / total_abas)
                self._progresso(prog_base, f"Importando {nome_aba}...")

                if nome_aba == "EMPRESA":
                    s, e = self._importar_empresa(wb, sobreescreve)
                elif nome_aba == "PRODUTOS":
                    s, e = self._importar_produtos_lote(wb, sobreescreve, prog_base, prog_range)
                elif nome_aba == "TRANSP":
                    s, e = self._importar_transportadoras_lote(wb, sobreescreve, prog_base, prog_range)
                elif nome_aba == "CLIENTES":
                    s, e = self._importar_clientes_lote(wb, sobreescreve, prog_base, prog_range)
                else:
                    s, e = self._importar_aba_padrao(wb, nome_aba, sobreescreve, prog_base, prog_range)

                resultados[nome_aba] = {"sucesso": s, "erros": e}

            elapsed_total = time.time() - t_total
            minutos, segundos = divmod(int(elapsed_total), 60)
            self._log(f"Tempo total: {minutos}m {segundos}s")
            self._progresso(100, "Importacao concluida!")
            wb.close()

        except Exception as e:
            self._log(f"ERRO FATAL: {e}")
            import traceback
            traceback.print_exc()
            resultados["ERRO_GERAL"] = str(e)
        finally:
            self.desconectar()

        return resultados
